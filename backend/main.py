from fastapi import FastAPI, HTTPException, UploadFile, File, Form, Request
from fastapi.responses import JSONResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from typing import List, Optional
from pathlib import Path
from uuid import uuid4
import time
import random
import re
import os
import json
import hmac
import hashlib
import base64
import csv
from urllib import request as urlrequest, error as urlerror
from openpyxl import load_workbook
from import_attendance_excel import import_attendance
from fastapi.middleware.cors import CORSMiddleware
from db import get_connection, init_db

app = FastAPI(title="Aviation ERP Lite")

# Allow frontend (local HTML/JS) to call backend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

BASE_DIR = Path(__file__).resolve().parent
FRONTEND_DIR = BASE_DIR.parent / "frontend"
EXCEL_PATH = BASE_DIR.parent / "attendance_master.xlsm"
SHEET_NAME = "attendance_log"
RECEIPTS_DIR = BASE_DIR / "receipts"
PASSWORDS_PATH = BASE_DIR.parent / "passwords.txt"
SESSION_TIMEOUT_SECONDS = 300
SESSIONS = {}
STAFF_USERS = {
    "praharsh": {"password": "9121726565", "display_name": "Praharsh", "welcome": "Welcome Praharsh Sir!"},
    "nanda": {"password": "8124326444", "display_name": "Nanda", "welcome": "Welcome Nanda Sir!"},
}
COURSE_FEES_INR = {
    "ground operations": 150000.0,
    "cabin crew": 250000.0,
}


def _load_local_env_file(path: Path):
    if not path.exists():
        return
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        k = key.strip()
        if not k:
            continue
        v = value.strip().strip('"').strip("'")
        os.environ.setdefault(k, v)


_load_local_env_file(BASE_DIR / ".env")

RAZORPAY_KEY_ID = os.getenv("RAZORPAY_KEY_ID", "").strip()
RAZORPAY_KEY_SECRET = os.getenv("RAZORPAY_KEY_SECRET", "").strip()
RESEND_API_KEY = os.getenv("RESEND_API_KEY", "").strip()
ADMISSIONS_TO_EMAIL = os.getenv("ADMISSIONS_TO_EMAIL", "thedanielraj@outlook.com").strip()
ADMISSIONS_FROM_EMAIL = os.getenv("ADMISSIONS_FROM_EMAIL", "Arunands ERP <onboarding@resend.dev>").strip()
ADMISSIONS_PDF_DIR = BASE_DIR / "admissions_pdfs"

init_db()

def _load_passwords():
    passwords = {}
    if PASSWORDS_PATH.exists():
        for line in PASSWORDS_PATH.read_text(encoding="utf-8").splitlines():
            if ":" in line:
                user, pwd = line.split(":", 1)
                passwords[user.strip()] = pwd.strip()
    return passwords

def _save_passwords(passwords: dict):
    lines = [f"{user}:{pwd}" for user, pwd in passwords.items()]
    PASSWORDS_PATH.write_text("\n".join(lines), encoding="utf-8")

def _generate_password():
    rng = random.SystemRandom()
    return "".join(str(rng.randint(0, 9)) for _ in range(8))

def _ensure_passwords_for_students():
    conn = get_connection()
    student_ids = [row["student_id"] for row in conn.execute("SELECT student_id FROM students").fetchall()]
    conn.close()

    passwords = _load_passwords()
    if "superuser" not in passwords:
        passwords["superuser"] = "qwerty"
    updated = False
    for username, config in STAFF_USERS.items():
        if username not in passwords:
            passwords[username] = str(config["password"])
            updated = True
    valid_ids = set()
    for sid in student_ids:
        if re.search(r"\d", str(sid)):
            valid_ids.add(sid)
            if sid not in passwords:
                passwords[sid] = _generate_password()
                updated = True

    # Remove invalid student ids (no digits) from passwords
    protected_users = {"superuser", *STAFF_USERS.keys()}
    to_remove = [u for u in passwords.keys() if u not in protected_users and u not in valid_ids]
    if to_remove:
        for u in to_remove:
            passwords.pop(u, None)
        updated = True

    if updated or not PASSWORDS_PATH.exists():
        _save_passwords(passwords)

_ensure_passwords_for_students()

class LoginRequest(BaseModel):
    username: str
    password: str


class TimetableEntryRequest(BaseModel):
    title: str
    day_of_week: str
    start_time: str
    end_time: str
    course: Optional[str] = ""
    batch: Optional[str] = ""
    location: Optional[str] = ""
    instructor: Optional[str] = ""


class InterviewStatRequest(BaseModel):
    airline_name: str
    interview_date: str
    notes: Optional[str] = ""


class AnnouncementRequest(BaseModel):
    title: str
    message: str


class NotificationRequest(BaseModel):
    title: str
    message: str
    level: Optional[str] = "info"
    target_user: Optional[str] = None


class RazorpayOrderRequest(BaseModel):
    student_id: Optional[str] = None
    amount_inr: Optional[float] = None


class RazorpayVerifyRequest(BaseModel):
    student_id: str
    razorpay_payment_id: str
    razorpay_order_id: str
    razorpay_signature: str
    amount_paid_inr: float


class FeePolicyUpdateRequest(BaseModel):
    student_id: str
    concession_amount: Optional[float] = 0
    due_date: Optional[str] = None


class ChangePasswordRequest(BaseModel):
    current_password: str
    new_password: str


class AdminSetPasswordRequest(BaseModel):
    username: str
    new_password: str


class AdmissionRequest(BaseModel):
    first_name: str
    middle_name: str = ""
    last_name: str
    phone: str
    email: str
    blood_group: str = ""
    age: int = 0
    dob: str = ""
    aadhaar_number: str = ""
    nationality: str = ""
    father_name: str = ""
    father_phone: str = ""
    father_occupation: str = ""
    father_email: str = ""
    mother_name: str = ""
    mother_phone: str = ""
    mother_occupation: str = ""
    mother_email: str = ""
    correspondence_address: str = ""
    permanent_address: str = ""
    course: str
    academic_details: List[dict] = []
    admission_pdf_base64: str = ""
    admission_pdf_filename: str = ""


class ActivityUndoRequest(BaseModel):
    activity_id: int


class StudentIdsRequest(BaseModel):
    student_ids: List[str]


class LeadRequest(BaseModel):
    name: str = ""
    age: str = ""
    qualification: str = ""
    location: str = ""
    phone: str
    preferred_time: str = ""
    intent: str = ""


class LeadFollowupRequest(BaseModel):
    followup_date: str


class TestQuestionRequest(BaseModel):
    question_text: str
    option_a: str
    option_b: str
    option_c: str
    option_d: str
    correct_answer: str


class TestCreateRequest(BaseModel):
    title: str
    description: str = ""
    duration_minutes: int = 30
    questions: List[TestQuestionRequest]
    assigned_students: List[str] = []


class TestSubmissionRequest(BaseModel):
    answers: List[dict]


class MalpracticeRequest(BaseModel):
    event_type: str
    details: str = ""


class ParentLinkRequest(BaseModel):
    student_id: str
    days: int = 30


class FeeRemindersRequest(BaseModel):
    days: int = 7


class StudentsBulkBatchRequest(BaseModel):
    student_ids: List[str]
    batch: str

def _create_session(username: str) -> str:
    token = uuid4().hex
    SESSIONS[token] = {"user": username, "last_activity": time.time()}
    return token

def _get_session(token: str):
    session = SESSIONS.get(token)
    if not session:
        return None
    if time.time() - session["last_activity"] > SESSION_TIMEOUT_SECONDS:
        SESSIONS.pop(token, None)
        return None
    session["last_activity"] = time.time()
    return session

def _get_current_user(request: Request):
    return getattr(request.state, "user", None)

def _is_superuser(user: dict) -> bool:
    if not user:
        return False
    username = str(user.get("user") or "").lower()
    return username == "superuser" or username in STAFF_USERS

def _require_superuser(user: dict):
    if not _is_superuser(user):
        raise HTTPException(status_code=403, detail="Forbidden")


def _require_self_or_superuser(user: dict, student_id: str):
    if user and not _is_superuser(user) and user["user"] != student_id:
        raise HTTPException(status_code=403, detail="Forbidden")


def _ensure_fee_policies_table():
    conn = get_connection()
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS fee_policies (
            student_id TEXT PRIMARY KEY,
            concession_amount REAL NOT NULL DEFAULT 0,
            due_date TEXT,
            updated_at TEXT NOT NULL DEFAULT (datetime('now'))
        )
        """
    )
    conn.commit()
    conn.close()


def _ensure_fee_receipt_columns(conn):
    cols = {row["name"] for row in conn.execute("PRAGMA table_info(fees)").fetchall()}
    additions = [
        ("payment_mode", "TEXT"),
        ("bank_name", "TEXT"),
        ("txn_utr_no", "TEXT"),
        ("bank_ref_no", "TEXT"),
        ("transaction_type", "TEXT"),
    ]
    for name, col_type in additions:
        if name not in cols:
            conn.execute(f"ALTER TABLE fees ADD COLUMN {name} {col_type}")
    conn.commit()


def _parse_razorpay_refs(remarks: str):
    text = str(remarks or "")
    payment_match = re.search(r"payment_id=([A-Za-z0-9_]+)", text, flags=re.IGNORECASE)
    order_match = re.search(r"order_id=([A-Za-z0-9_]+)", text, flags=re.IGNORECASE)
    return {
        "payment_id": payment_match.group(1) if payment_match else "",
        "order_id": order_match.group(1) if order_match else "",
    }


def _to_iso_date(value: str) -> str:
    raw = str(value or "").strip()
    if not raw:
        return time.strftime("%Y-%m-%d")
    if " " in raw and "T" not in raw:
        raw = raw.replace(" ", "T")
    if not raw.endswith("Z"):
        raw = f"{raw}Z"
    try:
        from datetime import datetime
        dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return time.strftime("%Y-%m-%d")


def _ensure_activity_table():
    conn = get_connection()
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS activity_log (
            activity_id INTEGER PRIMARY KEY AUTOINCREMENT,
            action_type TEXT NOT NULL,
            description TEXT NOT NULL,
            payload_json TEXT NOT NULL DEFAULT '{}',
            created_by TEXT NOT NULL,
            created_at TEXT NOT NULL DEFAULT (datetime('now')),
            undone INTEGER NOT NULL DEFAULT 0,
            undone_at TEXT
        )
        """
    )
    conn.commit()
    conn.close()


def _ensure_parent_links_table():
    conn = get_connection()
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS parent_links (
            token TEXT PRIMARY KEY,
            student_id TEXT NOT NULL,
            expires_at TEXT,
            created_at TEXT NOT NULL DEFAULT (datetime('now')),
            created_by TEXT
        )
        """
    )
    conn.commit()
    conn.close()


def _ensure_leads_table():
    conn = get_connection()
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS leads (
            lead_id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            age TEXT,
            qualification TEXT,
            location TEXT,
            phone TEXT,
            preferred_time TEXT,
            intent TEXT,
            source TEXT,
            last_message TEXT,
            last_reply TEXT,
            last_intent TEXT,
            updated_at TEXT,
            status TEXT NOT NULL DEFAULT 'new',
            contacted_at TEXT,
            followup_date TEXT,
            created_at TEXT NOT NULL DEFAULT (datetime('now'))
        )
        """
    )
    conn.commit()
    conn.close()


def _ensure_tests_tables():
    conn = get_connection()
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS tests (
            test_id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            description TEXT,
            duration_minutes INTEGER NOT NULL DEFAULT 30,
            is_active INTEGER NOT NULL DEFAULT 1,
            created_by TEXT NOT NULL,
            created_at TEXT NOT NULL DEFAULT (datetime('now'))
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS test_questions (
            question_id INTEGER PRIMARY KEY AUTOINCREMENT,
            test_id INTEGER NOT NULL,
            question_order INTEGER NOT NULL,
            question_text TEXT NOT NULL,
            question_type TEXT NOT NULL DEFAULT 'mcq',
            option_a TEXT,
            option_b TEXT,
            option_c TEXT,
            option_d TEXT,
            correct_answer TEXT,
            points INTEGER NOT NULL DEFAULT 1,
            FOREIGN KEY (test_id) REFERENCES tests(test_id) ON DELETE CASCADE
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS test_assignments (
            test_id INTEGER NOT NULL,
            student_id TEXT NOT NULL,
            assigned_at TEXT NOT NULL DEFAULT (datetime('now')),
            PRIMARY KEY (test_id, student_id)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS test_attempts (
            attempt_id INTEGER PRIMARY KEY AUTOINCREMENT,
            test_id INTEGER NOT NULL,
            student_id TEXT NOT NULL,
            start_time TEXT NOT NULL DEFAULT (datetime('now')),
            submitted_at TEXT,
            status TEXT NOT NULL DEFAULT 'in_progress',
            score REAL NOT NULL DEFAULT 0,
            total_points REAL NOT NULL DEFAULT 0,
            malpractice_count INTEGER NOT NULL DEFAULT 0,
            malpractice_flag INTEGER NOT NULL DEFAULT 0,
            question_order_json TEXT NOT NULL DEFAULT '[]',
            option_order_json TEXT NOT NULL DEFAULT '{}'
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS test_attempt_answers (
            answer_id INTEGER PRIMARY KEY AUTOINCREMENT,
            attempt_id INTEGER NOT NULL,
            question_id INTEGER NOT NULL,
            answer_text TEXT,
            is_correct INTEGER NOT NULL DEFAULT 0,
            points_awarded REAL NOT NULL DEFAULT 0
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS test_malpractice_events (
            event_id INTEGER PRIMARY KEY AUTOINCREMENT,
            attempt_id INTEGER NOT NULL,
            event_type TEXT NOT NULL,
            details TEXT,
            created_at TEXT NOT NULL DEFAULT (datetime('now'))
        )
        """
    )
    conn.commit()
    conn.close()


def _log_activity(user: dict, action_type: str, description: str, payload: dict):
    _ensure_activity_table()
    actor = (user or {}).get("user", "system")
    conn = get_connection()
    conn.execute(
        "INSERT INTO activity_log (action_type, description, payload_json, created_by) VALUES (?, ?, ?, ?)",
        (action_type, description, json.dumps(payload or {}), actor),
    )
    conn.commit()
    conn.close()


def _extract_airline_from_interview_remark(remark: str) -> str:
    text = (remark or "").strip()
    # Examples handled:
    # "Interview: Indigo", "Interview - Air India", "Indigo interview done"
    m = re.search(r"interview\s*[:\-]\s*([A-Za-z0-9 .&-]+)", text, re.IGNORECASE)
    if m:
        return m.group(1).strip()
    m = re.search(r"([A-Za-z][A-Za-z .&-]{2,})\s+interview", text, re.IGNORECASE)
    if m:
        return m.group(1).strip()
    return "Interview"


def _course_fee_inr(course: str) -> Optional[float]:
    if not course:
        return None
    return COURSE_FEES_INR.get(str(course).strip().lower())


def _student_financials(conn, student_id: str):
    _ensure_fee_policies_table()
    student = conn.execute(
        "SELECT student_id, student_name, course, batch FROM students WHERE student_id = ?",
        (student_id,),
    ).fetchone()
    if not student:
        return None

    fee = conn.execute(
        """
        SELECT
            COALESCE(SUM(amount_paid), 0) AS paid,
            COALESCE(MAX(amount_total), 0) AS max_total,
            COUNT(*) AS transactions
        FROM fees
        WHERE student_id = ?
        """,
        (student_id,),
    ).fetchone()

    policy = conn.execute(
        "SELECT concession_amount, due_date FROM fee_policies WHERE student_id = ?",
        (student_id,),
    ).fetchone()

    planned = _course_fee_inr(student["course"])
    base_total = float(planned if planned is not None else fee["max_total"])
    concession_amount = min(
        max(float((policy["concession_amount"] if policy else 0) or 0), 0.0),
        max(base_total, 0.0),
    )
    total = max(base_total - concession_amount, 0.0)
    paid = float(fee["paid"])
    due = max(total - paid, 0.0)
    return {
        "student": dict(student),
        "base_total": base_total,
        "concession_amount": concession_amount,
        "due_date": (policy["due_date"] if policy else None),
        "total": total,
        "paid": paid,
        "due": due,
        "transactions": int(fee["transactions"]),
    }


def _razorpay_auth_header() -> str:
    token = f"{RAZORPAY_KEY_ID}:{RAZORPAY_KEY_SECRET}".encode("utf-8")
    return "Basic " + base64.b64encode(token).decode("utf-8")


def _create_razorpay_order(amount_paise: int, receipt: str, notes: dict):
    payload = {
        "amount": amount_paise,
        "currency": "INR",
        "receipt": receipt,
        "notes": notes,
        "payment_capture": 1,
    }
    data = json.dumps(payload).encode("utf-8")
    req = urlrequest.Request(
        "https://api.razorpay.com/v1/orders",
        data=data,
        headers={
            "Authorization": _razorpay_auth_header(),
            "Content-Type": "application/json",
        },
        method="POST",
    )
    with urlrequest.urlopen(req, timeout=20) as resp:
        return json.loads(resp.read().decode("utf-8"))


def _send_admission_email(first_name: str, middle_name: str, last_name: str, phone: str, email: str, course: str, pdf_base64: str, pdf_filename: str):
    if not RESEND_API_KEY or not pdf_base64:
        return False
    full_name = " ".join([p for p in [first_name, middle_name, last_name] if p]).strip()
    payload = {
        "from": ADMISSIONS_FROM_EMAIL,
        "to": [ADMISSIONS_TO_EMAIL],
        "subject": f"New Admission - {full_name or 'Applicant'}",
        "text": "\n".join([
            "New admission form submitted.",
            f"Name: {full_name}",
            f"Course: {course}",
            f"Phone: {phone}",
            f"Email: {email}",
        ]),
        "attachments": [
            {
                "filename": pdf_filename or f"admission_{int(time.time())}.pdf",
                "content": pdf_base64,
            }
        ],
    }
    data = json.dumps(payload).encode("utf-8")
    req = urlrequest.Request(
        "https://api.resend.com/emails",
        data=data,
        headers={
            "Authorization": f"Bearer {RESEND_API_KEY}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    try:
        with urlrequest.urlopen(req, timeout=20):
            return True
    except Exception:
        return False


def _save_admission_pdf_local(pdf_base64: str, pdf_filename: str):
    encoded = str(pdf_base64 or "").strip()
    if not encoded:
        return {"saved": False, "path": None, "bytes": 0}
    try:
        raw = base64.b64decode(encoded, validate=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid admission PDF")
    max_bytes = 1024 * 1024
    if len(raw) > max_bytes:
        raise HTTPException(status_code=413, detail="Admission PDF must be less than 1 MB")
    ADMISSIONS_PDF_DIR.mkdir(parents=True, exist_ok=True)
    safe_name = re.sub(r"[^A-Za-z0-9._-]", "_", (pdf_filename or "").strip()) or f"admission_{int(time.time())}.pdf"
    if not safe_name.lower().endswith(".pdf"):
        safe_name += ".pdf"
    filepath = ADMISSIONS_PDF_DIR / f"{uuid4().hex}_{safe_name}"
    filepath.write_bytes(raw)
    return {"saved": True, "path": str(filepath), "bytes": len(raw)}

@app.middleware("http")
async def auth_middleware(request: Request, call_next):
    path = request.url.path
    if request.method == "OPTIONS":
        return await call_next(request)
    if path in ["/", "/login", "/auth/me", "/public/student-ids", "/public/alumni", "/admissions/apply", "/docs", "/openapi.json", "/style.css"] or path.startswith("/static") or path.startswith("/js/") or path.startswith("/assets/"):
        return await call_next(request)
    auth = request.headers.get("Authorization", "")
    if not auth.startswith("Bearer "):
        return JSONResponse(status_code=401, content={"detail": "Unauthorized"})
    token = auth.replace("Bearer ", "").strip()
    session = _get_session(token)
    if not session:
        return JSONResponse(status_code=401, content={"detail": "Session expired"})
    request.state.user = session
    return await call_next(request)

class AttendanceRecord(BaseModel):
    student_id: str
    student_name: str
    course: str
    batch: str
    attendance_status: str
    remarks: str = ""

class AttendanceSubmission(BaseModel):
    date: str
    records: List[AttendanceRecord]

def format_date_ddmmyyyy(date_str: str) -> str:
    # Expecting YYYY-MM-DD from the UI
    parts = date_str.split("-")
    if len(parts) == 3:
        yyyy, mm, dd = parts
        return f"{dd}-{mm}-{yyyy}"
    return date_str

def to_excel_status(status: str) -> str:
    return "P" if status.strip().lower() == "present" else "A"


def normalize_date_to_iso(raw: str) -> str:
    s = str(raw or "").strip()
    if not s:
        return ""
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        return s
    if re.match(r"^\d{2}-\d{2}-\d{4}$", s):
        d, m, y = s.split("-")
        return f"{y}-{m}-{d}"
    if re.match(r"^\d{8}$", s):
        d, m, y = s[0:2], s[2:4], s[4:8]
        return f"{y}-{m}-{d}"
    return s


def normalize_attendance_status(raw: str) -> str:
    s = str(raw or "").strip().lower()
    if s in ("p", "present"):
        return "Present"
    if s in ("a", "absent"):
        return "Absent"
    if not s:
        return "Absent"
    return s[0].upper() + s[1:]

@app.get("/")
def home():
    return FileResponse(FRONTEND_DIR / "index.html")

app.mount("/static", StaticFiles(directory=FRONTEND_DIR), name="static")
app.mount("/js", StaticFiles(directory=FRONTEND_DIR / "js"), name="js")
app.mount("/assets", StaticFiles(directory=FRONTEND_DIR / "assets"), name="assets")


@app.get("/style.css")
def style_file():
    return FileResponse(FRONTEND_DIR / "style.css")


@app.get("/public/student-ids")
def public_student_ids():
    _ensure_passwords_for_students()
    passwords = _load_passwords()
    ids = [u for u in passwords.keys() if u != "superuser" and str(u).upper().startswith("AAI")]
    ids.sort(reverse=True)
    return ids


@app.get("/public/alumni")
def public_alumni():
    conn = get_connection()
    selected_rows = conn.execute(
        """
        SELECT student_id, student_name, MAX(date) AS last_selected_date
        FROM attendance
        WHERE remarks IS NOT NULL
          AND TRIM(remarks) <> ''
          AND lower(remarks) LIKE '%selected%'
        GROUP BY student_id, student_name
        """
    ).fetchall()
    alumni_rows = conn.execute(
        """
        SELECT student_id, student_name, NULL AS last_selected_date
        FROM students
        WHERE lower(COALESCE(status, 'active')) = 'alumni'
        """
    ).fetchall()
    conn.close()
    by_id = {}
    for row in list(selected_rows) + list(alumni_rows):
        by_id[str(row["student_id"])] = dict(row)
    rows = list(by_id.values())
    rows.sort(key=lambda r: str(r.get("student_name") or "").lower())
    rows.sort(key=lambda r: str(r.get("last_selected_date") or ""), reverse=True)
    return rows


def _ensure_admissions_table():
    conn = get_connection()
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS admissions (
            admission_id INTEGER PRIMARY KEY AUTOINCREMENT,
            full_name TEXT NOT NULL DEFAULT '',
            first_name TEXT NOT NULL,
            middle_name TEXT,
            last_name TEXT NOT NULL,
            phone TEXT NOT NULL,
            email TEXT NOT NULL,
            blood_group TEXT,
            age INTEGER,
            dob TEXT,
            aadhaar_number TEXT,
            nationality TEXT,
            father_name TEXT,
            father_phone TEXT,
            father_occupation TEXT,
            father_email TEXT,
            mother_name TEXT,
            mother_phone TEXT,
            mother_occupation TEXT,
            mother_email TEXT,
            correspondence_address TEXT,
            permanent_address TEXT,
            course TEXT NOT NULL,
            academic_details_json TEXT NOT NULL DEFAULT '[]',
            admission_pdf_path TEXT,
            admission_pdf_bytes INTEGER,
            status TEXT NOT NULL DEFAULT 'new',
            created_at TEXT NOT NULL DEFAULT (datetime('now'))
        )
        """
    )
    existing = {row["name"] for row in conn.execute("PRAGMA table_info(admissions)").fetchall()}
    expected = {
        "full_name": "TEXT NOT NULL DEFAULT ''",
        "first_name": "TEXT",
        "middle_name": "TEXT",
        "last_name": "TEXT",
        "phone": "TEXT",
        "email": "TEXT",
        "blood_group": "TEXT",
        "age": "INTEGER",
        "dob": "TEXT",
        "aadhaar_number": "TEXT",
        "nationality": "TEXT",
        "father_name": "TEXT",
        "father_phone": "TEXT",
        "father_occupation": "TEXT",
        "father_email": "TEXT",
        "mother_name": "TEXT",
        "mother_phone": "TEXT",
        "mother_occupation": "TEXT",
        "mother_email": "TEXT",
        "correspondence_address": "TEXT",
        "permanent_address": "TEXT",
        "course": "TEXT",
        "academic_details_json": "TEXT NOT NULL DEFAULT '[]'",
        "admission_pdf_path": "TEXT",
        "admission_pdf_bytes": "INTEGER",
        "status": "TEXT",
        "created_at": "TEXT",
    }
    for col, typ in expected.items():
        if col not in existing:
            conn.execute(f"ALTER TABLE admissions ADD COLUMN {col} {typ}")
    conn.commit()
    conn.close()


@app.post("/admissions/apply")
def admissions_apply(payload: AdmissionRequest):
    first_name = payload.first_name.strip()
    middle_name = payload.middle_name.strip()
    last_name = payload.last_name.strip()
    full_name = " ".join([p for p in [first_name, middle_name, last_name] if p]).strip()
    phone = payload.phone.strip()
    email = payload.email.strip()
    blood_group = payload.blood_group.strip()
    age = int(payload.age or 0)
    dob = payload.dob.strip()
    aadhaar_number = payload.aadhaar_number.strip()
    nationality = payload.nationality.strip()
    father_name = payload.father_name.strip()
    father_phone = payload.father_phone.strip()
    father_occupation = payload.father_occupation.strip()
    father_email = payload.father_email.strip()
    mother_name = payload.mother_name.strip()
    mother_phone = payload.mother_phone.strip()
    mother_occupation = payload.mother_occupation.strip()
    mother_email = payload.mother_email.strip()
    correspondence_address = payload.correspondence_address.strip()
    permanent_address = payload.permanent_address.strip()
    course = payload.course.strip()
    academic_details_json = json.dumps(payload.academic_details or [])
    admission_pdf_base64 = (payload.admission_pdf_base64 or "").strip()
    admission_pdf_filename = (payload.admission_pdf_filename or "").strip()
    if not first_name or not last_name or not phone or not email or not course:
        raise HTTPException(status_code=400, detail="Missing required fields")

    _ensure_admissions_table()
    pdf_saved = _save_admission_pdf_local(admission_pdf_base64, admission_pdf_filename)
    conn = get_connection()
    cur = conn.execute(
        """
        INSERT INTO admissions (
            full_name,
            first_name, middle_name, last_name, phone, email, blood_group, age, dob, aadhaar_number, nationality,
            father_name, father_phone, father_occupation, father_email, mother_name, mother_phone, mother_occupation, mother_email,
            correspondence_address, permanent_address, course, academic_details_json, admission_pdf_path, admission_pdf_bytes
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            full_name,
            first_name, middle_name, last_name, phone, email, blood_group, age, dob, aadhaar_number, nationality,
            father_name, father_phone, father_occupation, father_email, mother_name, mother_phone, mother_occupation, mother_email,
            correspondence_address, permanent_address, course, academic_details_json,
            pdf_saved["path"],
            int(pdf_saved["bytes"]),
        ),
    )
    admission_id = cur.lastrowid
    conn.commit()
    conn.close()
    _log_activity(
        {"user": "public_admission_form"},
        "admission_submitted",
        f"Admission submitted: {first_name} {last_name} ({course})",
        {
            "admission_id": admission_id,
            "first_name": first_name,
            "last_name": last_name,
            "course": course,
            "phone": phone,
            "email": email,
            "admission_pdf_path": pdf_saved["path"],
            "admission_pdf_bytes": int(pdf_saved["bytes"]),
        },
    )
    email_sent = _send_admission_email(
        first_name, middle_name, last_name, phone, email, course, admission_pdf_base64, admission_pdf_filename
    )
    return {
        "status": "ok",
        "message": "Admission form submitted",
        "email_sent": email_sent,
        "pdf_stored": bool(pdf_saved["saved"]),
        "pdf_bytes": int(pdf_saved["bytes"]),
    }


@app.get("/admissions")
def admissions_list(request: Request):
    user = _get_current_user(request)
    _require_superuser(user)
    _ensure_admissions_table()
    conn = get_connection()
    rows = conn.execute(
        """
        SELECT
            admission_id,
            first_name,
            middle_name,
            last_name,
            course,
            phone,
            email,
            created_at,
            admission_pdf_path,
            admission_pdf_bytes
        FROM admissions
        ORDER BY admission_id DESC
        LIMIT 500
        """
    ).fetchall()
    conn.close()
    out = []
    for r in rows:
        full_name = " ".join([p for p in [r["first_name"], r["middle_name"], r["last_name"]] if p]).strip()
        out.append(
            {
                "admission_id": r["admission_id"],
                "full_name": full_name,
                "course": r["course"] or "",
                "phone": r["phone"] or "",
                "email": r["email"] or "",
                "created_at": r["created_at"] or "",
                "pdf_available": bool((r["admission_pdf_path"] or "").strip()),
                "pdf_bytes": int(r["admission_pdf_bytes"] or 0),
            }
        )
    return out


@app.post("/leads")
def create_lead(payload: LeadRequest):
    _ensure_leads_table()
    conn = get_connection()
    conn.execute(
        """
        INSERT INTO leads (name, age, qualification, location, phone, preferred_time, intent, source, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, 'inquiry', datetime('now'))
        """,
        (payload.name, payload.age, payload.qualification, payload.location, payload.phone, payload.preferred_time, payload.intent),
    )
    conn.commit()
    conn.close()
    return {"status": "ok", "message": "Lead captured"}


@app.get("/leads")
def list_leads(request: Request):
    user = _get_current_user(request)
    _require_superuser(user)
    _ensure_leads_table()
    conn = get_connection()
    rows = conn.execute("SELECT * FROM leads ORDER BY lead_id DESC").fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.post("/leads/{lead_id}/contacted")
def mark_lead_contacted(lead_id: int, request: Request):
    user = _get_current_user(request)
    _require_superuser(user)
    _ensure_leads_table()
    conn = get_connection()
    conn.execute(
        "UPDATE leads SET status = 'contacted', contacted_at = datetime('now') WHERE lead_id = ?",
        (lead_id,),
    )
    conn.commit()
    conn.close()
    return {"status": "ok"}


@app.post("/leads/{lead_id}/not-interested")
def mark_lead_not_interested(lead_id: int, request: Request):
    user = _get_current_user(request)
    _require_superuser(user)
    _ensure_leads_table()
    conn = get_connection()
    conn.execute(
        "UPDATE leads SET status = 'not_interested' WHERE lead_id = ?",
        (lead_id,),
    )
    conn.commit()
    conn.close()
    return {"status": "ok"}


@app.post("/leads/{lead_id}/followup")
def set_lead_followup(lead_id: int, payload: LeadFollowupRequest, request: Request):
    user = _get_current_user(request)
    _require_superuser(user)
    _ensure_leads_table()
    conn = get_connection()
    conn.execute(
        "UPDATE leads SET followup_date = ? WHERE lead_id = ?",
        (payload.followup_date, lead_id),
    )
    conn.commit()
    conn.close()
    return {"status": "ok"}


@app.post("/parent/link")
def create_parent_link(payload: ParentLinkRequest, request: Request):
    user = _get_current_user(request)
    _require_superuser(user)
    _ensure_parent_links_table()
    token = uuid4().hex
    expires_at = (Path(__file__).resolve().parent.parent / "frontend").exists() # Just a dummy check for now
    from datetime import datetime, timedelta
    expires_dt = datetime.now() + timedelta(days=payload.days)
    expires_at_str = expires_dt.strftime("%Y-%m-%d")

    conn = get_connection()
    conn.execute(
        "INSERT INTO parent_links (token, student_id, expires_at, created_by) VALUES (?, ?, ?, ?)",
        (token, payload.student_id, expires_at_str, user["user"]),
    )
    conn.commit()
    conn.close()
    return {"status": "ok", "token": token, "student_id": payload.student_id, "expires_at": expires_at_str}


@app.get("/parent/summary")
def parent_summary(token: str):
    _ensure_parent_links_table()
    conn = get_connection()
    link = conn.execute("SELECT student_id, expires_at FROM parent_links WHERE token = ?", (token,)).fetchone()
    if not link:
        conn.close()
        raise HTTPException(status_code=404, detail="Link not found")

    if link["expires_at"] and link["expires_at"] < time.strftime("%Y-%m-%d"):
        conn.close()
        raise HTTPException(status_code=410, detail="Link expired")

    student_id = link["student_id"]
    info = _student_financials(conn, student_id)
    if not info:
        conn.close()
        raise HTTPException(status_code=404, detail="Student not found")

    attendance = conn.execute(
        "SELECT date, attendance_status, remarks FROM attendance WHERE student_id = ? ORDER BY date DESC LIMIT 10",
        (student_id,),
    ).fetchall()
    conn.close()

    return {
        "student": info["student"],
        "fees": {
            "total": info["total"],
            "paid": info["paid"],
            "due": info["due"],
            "due_date": info["due_date"]
        },
        "attendance": [dict(r) for r in attendance]
    }


@app.get("/admissions/{admission_id}/pdf")
def admissions_pdf(admission_id: int, request: Request):
    user = _get_current_user(request)
    _require_superuser(user)
    _ensure_admissions_table()
    conn = get_connection()
    row = conn.execute(
        "SELECT admission_pdf_path, first_name, last_name FROM admissions WHERE admission_id = ?",
        (admission_id,),
    ).fetchone()
    conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="Admission not found")
    pdf_path = str(row["admission_pdf_path"] or "").strip()
    if not pdf_path:
        raise HTTPException(status_code=404, detail="Admission PDF not found")
    path = Path(pdf_path)
    if not path.exists() or not path.is_file():
        raise HTTPException(status_code=404, detail="Admission PDF file missing")
    safe_first = re.sub(r"[^A-Za-z0-9._-]", "_", str(row["first_name"] or "admission"))
    safe_last = re.sub(r"[^A-Za-z0-9._-]", "_", str(row["last_name"] or ""))
    filename = f"{safe_first}{'_' + safe_last if safe_last else ''}_{admission_id}.pdf"
    return FileResponse(path, media_type="application/pdf", filename=filename)


@app.delete("/admissions/{admission_id}")
def admissions_delete(admission_id: int, request: Request):
    user = _get_current_user(request)
    _require_superuser(user)
    _ensure_admissions_table()
    conn = get_connection()
    row = conn.execute("SELECT * FROM admissions WHERE admission_id = ?", (admission_id,)).fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Admission not found")
    conn.execute("DELETE FROM admissions WHERE admission_id = ?", (admission_id,))
    conn.commit()
    conn.close()
    full_name = str(row["full_name"] or " ".join([p for p in [row["first_name"], row["middle_name"], row["last_name"]] if p]).strip())
    _log_activity(
        user,
        "admission_deleted",
        f"Deleted admission #{admission_id}{f' ({full_name})' if full_name else ''}",
        {"admission": dict(row)},
    )
    return {"status": "ok", "message": "Admission deleted"}

@app.get("/students")
def list_students(request: Request):
    user = _get_current_user(request)
    conn = get_connection()
    if user and not _is_superuser(user):
        rows = conn.execute(
            "SELECT * FROM students WHERE student_id = ? ORDER BY student_id DESC",
            (user["user"],),
        ).fetchall()
    else:
        rows = conn.execute("SELECT * FROM students ORDER BY student_id DESC").fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/students")
def add_student(student_name: str, course: str, batch: str, request: Request):
    user = _get_current_user(request)
    _require_superuser(user)
    conn = get_connection()
    cur = conn.execute(
        "INSERT INTO students (student_name, course, batch) VALUES (?, ?, ?)",
        (student_name, course, batch),
    )
    sid = str(cur.lastrowid)
    conn.commit()
    conn.close()
    _log_activity(user, "student_added", f"Student added: {student_name} ({sid})", {
        "student_id": sid,
        "student_name": student_name,
        "course": course,
        "batch": batch,
    })
    _ensure_passwords_for_students()
    return {"status": "ok", "message": "Student added"}


def _normalize_student_ids(values: List[str]) -> List[str]:
    seen = set()
    cleaned = []
    for value in values or []:
        sid = str(value or "").strip()
        if sid and sid not in seen:
            seen.add(sid)
            cleaned.append(sid)
    return cleaned


@app.post("/students/bulk-batch")
def students_bulk_batch(payload: StudentsBulkBatchRequest, request: Request):
    user = _get_current_user(request)
    _require_superuser(user)
    student_ids = _normalize_student_ids(payload.student_ids)
    batch = str(payload.batch or "").strip()
    if not student_ids or not batch:
        raise HTTPException(status_code=400, detail="Invalid payload")
    conn = get_connection()
    previous_batches = {}
    for sid in student_ids:
        row = conn.execute("SELECT batch FROM students WHERE student_id = ?", (sid,)).fetchone()
        if row and row["batch"]:
            previous_batches[sid] = str(row["batch"])
        conn.execute("UPDATE students SET batch = ? WHERE student_id = ?", (batch, sid))
        conn.execute("UPDATE attendance SET batch = ? WHERE student_id = ?", (batch, sid))
    conn.commit()
    conn.close()
    _log_activity(
        user,
        "students_batch_updated",
        f"Moved {len(student_ids)} student(s) to batch {batch}",
        {"student_ids": student_ids, "batch": batch, "previous_batches": previous_batches},
    )
    return {"status": "ok", "updated": len(student_ids)}


@app.post("/students/mark-alumni")
def students_mark_alumni(payload: StudentIdsRequest, request: Request):
    user = _get_current_user(request)
    _require_superuser(user)
    student_ids = _normalize_student_ids(payload.student_ids)
    if not student_ids:
        raise HTTPException(status_code=400, detail="Invalid payload")
    conn = get_connection()
    previous_status = {}
    for sid in student_ids:
        row = conn.execute("SELECT status FROM students WHERE student_id = ?", (sid,)).fetchone()
        previous_status[sid] = str((row["status"] if row else "Active") or "Active")
        conn.execute("UPDATE students SET status = 'Alumni' WHERE student_id = ?", (sid,))
    conn.commit()
    conn.close()
    _log_activity(
        user,
        "students_marked_alumni",
        f"Marked {len(student_ids)} student(s) as alumni",
        {"student_ids": student_ids, "previous_status": previous_status},
    )
    return {"status": "ok", "updated": len(student_ids)}


@app.post("/students/delete")
def students_delete(payload: StudentIdsRequest, request: Request):
    user = _get_current_user(request)
    _require_superuser(user)
    student_ids = _normalize_student_ids(payload.student_ids)
    if not student_ids:
        raise HTTPException(status_code=400, detail="Invalid payload")
    conn = get_connection()
    snapshots = []
    for sid in student_ids:
        student = conn.execute(
            "SELECT student_id, student_name, course, batch, status FROM students WHERE student_id = ?",
            (sid,),
        ).fetchone()
        if not student:
            continue
        attendance_rows = conn.execute(
            "SELECT student_name, course, batch, date, attendance_status, remarks FROM attendance WHERE student_id = ?",
            (sid,),
        ).fetchall()
        fee_rows = conn.execute(
            "SELECT amount_total, amount_paid, due_date, remarks, receipt_path FROM fees WHERE student_id = ?",
            (sid,),
        ).fetchall()
        snapshots.append({
            "student": dict(student),
            "attendance": [dict(r) for r in attendance_rows],
            "fees": [dict(r) for r in fee_rows],
        })
        conn.execute("DELETE FROM attendance WHERE student_id = ?", (sid,))
        conn.execute("DELETE FROM fees WHERE student_id = ?", (sid,))
        conn.execute("DELETE FROM students WHERE student_id = ?", (sid,))
    conn.commit()
    conn.close()
    passwords = _load_passwords()
    changed = False
    for sid in student_ids:
        if sid in passwords:
            passwords.pop(sid, None)
            changed = True
    if changed:
        _save_passwords(passwords)
    _log_activity(
        user,
        "students_deleted",
        f"Deleted {len(snapshots)} student(s)",
        {"students": snapshots},
    )
    return {"status": "ok", "deleted": len(snapshots)}

@app.get("/students/{student_id}/balance")
def student_balance(student_id: str, request: Request):
    user = _get_current_user(request)
    _require_self_or_superuser(user, student_id)
    conn = get_connection()
    info = _student_financials(conn, student_id)
    conn.close()
    if not info:
        raise HTTPException(status_code=404, detail="Student not found")

    return {
        "student_id": student_id,
        "student_name": info["student"]["student_name"],
        "course": info["student"]["course"],
        "total": info["total"],
        "paid": info["paid"],
        "balance": info["due"],
        "concession_amount": info["concession_amount"],
        "due_date": info["due_date"],
        "gst_percent": 18,
    }

@app.get("/students/{student_id}/attendance")
def student_attendance(student_id: str, request: Request):
    user = _get_current_user(request)
    _require_self_or_superuser(user, student_id)
    conn = get_connection()
    rows = conn.execute("""
        SELECT date, attendance_status, remarks
        FROM attendance
        WHERE student_id = ?
        ORDER BY date DESC
    """, (student_id,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.get("/students/{student_id}/password")
def student_password(student_id: str, request: Request):
    user = _get_current_user(request)
    _require_superuser(user)
    _ensure_passwords_for_students()
    conn = get_connection()
    exists = conn.execute("SELECT 1 FROM students WHERE student_id = ?", (student_id,)).fetchone()
    conn.close()
    if not exists:
        raise HTTPException(status_code=404, detail="Student not found")
    passwords = _load_passwords()
    if student_id not in passwords:
        passwords[student_id] = _generate_password()
        _save_passwords(passwords)
    return {"student_id": student_id, "password": passwords[student_id]}

@app.get("/students/{student_id}/fees")
def student_fees(student_id: str, request: Request):
    user = _get_current_user(request)
    _require_self_or_superuser(user, student_id)
    conn = get_connection()
    rows = conn.execute("""
        SELECT fee_id, amount_total, amount_paid, due_date, remarks, created_at
        FROM fees
        WHERE student_id = ?
        ORDER BY fee_id DESC
    """, (student_id,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/fees/pay")
def add_payment(student_id: str, amount_total: float, amount_paid: float, remarks: str = "", request: Request = None):
    _require_superuser(_get_current_user(request))
    conn = get_connection()
    conn.execute("""
        INSERT INTO fees (student_id, amount_total, amount_paid, remarks)
        VALUES (?, ?, ?, ?)
    """, (student_id, amount_total, amount_paid, remarks))
    conn.commit()
    conn.close()
    return {"status": "ok", "message": "Fee entry added"}

@app.get("/students/{student_id}/profile")
def get_student_profile(student_id: str, request: Request):
    user = _get_current_user(request)
    _require_self_or_superuser(user, student_id)
    
    conn = get_connection()
    
    # Get profile data
    profile_row = conn.execute("""
        SELECT student_phone, student_email, aadhaar_number, pan_number, blood_group, religion, mother_tongue, address_details,
               parent_name, parent_occupation, parent_aadhaar, parent_qualification, parent_office_address, parent_office_phone, parent_email, parent_address,
               guardian_name, guardian_relation, guardian_phone, guardian_aadhaar, guardian_email, guardian_address
        FROM student_profiles
        WHERE student_id = ?
    """, (student_id,)).fetchone()
    
    profile_data = {}
    if profile_row:
        profile_data = dict(profile_row)
    
    # Get files
    files_rows = conn.execute("""
        SELECT file_type, file_name, mime_type, length(file_data) as file_size
        FROM profile_files
        WHERE student_id = ?
        ORDER BY uploaded_at DESC
    """, (student_id,)).fetchall()
    
    files = {}
    for row in files_rows:
        files[row['file_type']] = {
            'available': True,
            'bytes': row['file_size'],
            'filename': row['file_name'],
            'mime_type': row['mime_type']
        }
    
    conn.close()
    
    return {**profile_data, 'files': files}

@app.post("/students/{student_id}/profile")
def save_student_profile(
    student_id: str,
    request: Request,
    student_phone: str = Form(None),
    student_email: str = Form(None),
    aadhaar_number: str = Form(None),
    pan_number: str = Form(None),
    blood_group: str = Form(None),
    religion: str = Form(None),
    mother_tongue: str = Form(None),
    address_details: str = Form(None),
    parent_name: str = Form(None),
    parent_occupation: str = Form(None),
    parent_aadhaar: str = Form(None),
    parent_qualification: str = Form(None),
    parent_office_address: str = Form(None),
    parent_office_phone: str = Form(None),
    parent_email: str = Form(None),
    parent_address: str = Form(None),
    guardian_name: str = Form(None),
    guardian_relation: str = Form(None),
    guardian_phone: str = Form(None),
    guardian_aadhaar: str = Form(None),
    guardian_email: str = Form(None),
    guardian_address: str = Form(None),
    student_photo_base64: str = Form(None),
    student_photo_filename: str = Form(None),
    student_photo_type: str = Form(None),
    parent_photo_base64: str = Form(None),
    parent_photo_filename: str = Form(None),
    parent_photo_type: str = Form(None),
    guardian_photo_base64: str = Form(None),
    guardian_photo_filename: str = Form(None),
    guardian_photo_type: str = Form(None),
    admission_form_base64: str = Form(None),
    admission_form_filename: str = Form(None),
    admission_form_type: str = Form(None),
    pan_card_base64: str = Form(None),
    pan_card_filename: str = Form(None),
    pan_card_type: str = Form(None),
    aadhaar_card_base64: str = Form(None),
    aadhaar_card_filename: str = Form(None),
    aadhaar_card_type: str = Form(None)
):
    user = _get_current_user(request)
    _require_self_or_superuser(user, student_id)
    
    conn = get_connection()
    
    # Upsert profile data
    conn.execute("""
        INSERT OR REPLACE INTO student_profiles 
        (student_id, student_phone, student_email, aadhaar_number, pan_number, blood_group, religion, mother_tongue, address_details,
         parent_name, parent_occupation, parent_aadhaar, parent_qualification, parent_office_address, parent_office_phone, parent_email, parent_address,
         guardian_name, guardian_relation, guardian_phone, guardian_aadhaar, guardian_email, guardian_address, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))
    """, (
        student_id, student_phone, student_email, aadhaar_number, pan_number, blood_group, religion, mother_tongue, address_details,
        parent_name, parent_occupation, parent_aadhaar, parent_qualification, parent_office_address, parent_office_phone, parent_email, parent_address,
        guardian_name, guardian_relation, guardian_phone, guardian_aadhaar, guardian_email, guardian_address
    ))
    
    # Handle file uploads
    file_fields = [
        ('student_photo', student_photo_base64, student_photo_filename, student_photo_type),
        ('parent_photo', parent_photo_base64, parent_photo_filename, parent_photo_type),
        ('guardian_photo', guardian_photo_base64, guardian_photo_filename, guardian_photo_type),
        ('admission_form', admission_form_base64, admission_form_filename, admission_form_type),
        ('pan_card', pan_card_base64, pan_card_filename, pan_card_type),
        ('aadhaar_card', aadhaar_card_base64, aadhaar_card_filename, aadhaar_card_type)
    ]
    
    for file_type, base64_data, filename, mime_type in file_fields:
        if base64_data:
            # Decode base64 to bytes
            file_data = base64.b64decode(base64_data)
            
            # Delete existing file of this type
            conn.execute("DELETE FROM profile_files WHERE student_id = ? AND file_type = ?", (student_id, file_type))
            
            # Insert new file
            conn.execute("""
                INSERT INTO profile_files (student_id, file_type, file_name, file_data, mime_type)
                VALUES (?, ?, ?, ?, ?)
            """, (student_id, file_type, filename, file_data, mime_type))
    
    conn.commit()
    conn.close()
    
    return {"status": "ok", "message": "Profile updated successfully"}

@app.get("/students/{student_id}/profile/files/{file_type}")
def get_student_profile_file(student_id: str, file_type: str, request: Request):
    user = _get_current_user(request)
    _require_self_or_superuser(user, student_id)
    
    conn = get_connection()
    row = conn.execute("""
        SELECT file_name, file_data, mime_type
        FROM profile_files
        WHERE student_id = ? AND file_type = ?
        ORDER BY uploaded_at DESC
        LIMIT 1
    """, (student_id, file_type)).fetchone()
    conn.close()
    
    if not row:
        raise HTTPException(status_code=404, detail="File not found")
    
    return FileResponse(
        content=row['file_data'],
        media_type=row['mime_type'] or 'application/octet-stream',
        filename=row['file_name'] or f"{file_type}.bin"
    )

@app.get("/attendance/recent")
def recent_attendance(request: Request):
    user = _get_current_user(request)
    conn = get_connection()
    if user and not _is_superuser(user):
        rows = conn.execute(
            """
            SELECT student_id, student_name, date, attendance_status
            FROM attendance
            WHERE student_id = ?
            ORDER BY date DESC
            LIMIT 20
            """,
            (user["user"],),
        ).fetchall()
    else:
        rows = conn.execute(
            """
            SELECT student_id, student_name, date, attendance_status
            FROM attendance
            ORDER BY date DESC
            LIMIT 20
            """
        ).fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.get("/attendance/by-date")
def attendance_by_date(date: str, request: Request):
    user = _get_current_user(request)
    conn = get_connection()
    if user and not _is_superuser(user):
        rows = conn.execute(
            """
            SELECT student_id, student_name, date, attendance_status, remarks
            FROM attendance
            WHERE date = ? AND student_id = ?
            ORDER BY student_name ASC
            """,
            (date, user["user"]),
        ).fetchall()
    else:
        rows = conn.execute(
            """
            SELECT student_id, student_name, date, attendance_status, remarks
            FROM attendance
            WHERE date = ?
            ORDER BY student_name ASC
            """,
            (date,),
        ).fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.get("/attendance/month")
def attendance_month(month: str, request: Request):
    user = _get_current_user(request)
    if not re.match(r"^\d{4}-\d{2}$", month):
        raise HTTPException(status_code=400, detail="Invalid month")
    year, mon = map(int, month.split("-"))
    start = f"{year}-{mon:02d}-01"
    if mon == 12:
        end = f"{year+1}-01-01"
    else:
        end = f"{year}-{mon+1:02d}-01"

    conn = get_connection()
    if user and not _is_superuser(user):
        rows = conn.execute(
            "SELECT date, attendance_status AS status FROM attendance WHERE student_id = ? AND date >= ? AND date < ?",
            (user["user"], start, end),
        ).fetchall()
        mode = "student"
    else:
        rows = conn.execute(
            """
            SELECT date,
                   SUM(CASE WHEN attendance_status = 'Present' THEN 1 ELSE 0 END) AS present,
                   SUM(CASE WHEN attendance_status = 'Absent' THEN 1 ELSE 0 END) AS absent
            FROM attendance
            WHERE date >= ? AND date < ?
            GROUP BY date
            """,
            (start, end),
        ).fetchall()
        mode = "staff"
    conn.close()
    return {"mode": mode, "month": month, "days": [dict(r) for r in rows]}


@app.post("/attendance/record")
def record_attendance(payload: AttendanceSubmission, request: Request):
    user = _get_current_user(request)
    _require_superuser(user)
    if not payload.records:
        raise HTTPException(status_code=400, detail="No attendance records provided")
    today_iso = time.strftime("%Y-%m-%d")
    if str(payload.date) > today_iso:
        raise HTTPException(status_code=400, detail="Future attendance dates are not allowed")

    # Append to Excel
    if not EXCEL_PATH.exists():
        raise HTTPException(status_code=500, detail="attendance_master.xlsm not found")

    wb = load_workbook(EXCEL_PATH, keep_vba=True)
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
    else:
        ws = wb.active

    if ws.max_row == 1 and ws.max_column == 1 and ws.cell(1, 1).value is None:
        ws.append(["student_id", "student_name", "course", "batch", "date", "attendance_status", "remarks"])

    excel_date = format_date_ddmmyyyy(payload.date)
    for r in payload.records:
        ws.append([
            r.student_id,
            r.student_name,
            r.course,
            r.batch,
            excel_date,
            to_excel_status(r.attendance_status),
            r.remarks
        ])

    wb.save(EXCEL_PATH)

    # Insert into DB (avoid duplicates by UNIQUE constraint)
    conn = get_connection()
    inserted_ids = []
    for r in payload.records:
        conn.execute(
            """
            INSERT OR IGNORE INTO students
            (student_id, student_name, course, batch)
            VALUES (?, ?, ?, ?)
            """,
            (r.student_id, r.student_name, r.course, r.batch),
        )
        cur = conn.execute(
            """
            INSERT OR IGNORE INTO attendance
            (student_id, student_name, course, batch, date, attendance_status, remarks)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (r.student_id, r.student_name, r.course, r.batch, payload.date, r.attendance_status, r.remarks),
        )
        if cur.rowcount:
            inserted_ids.append(str(r.student_id))
    conn.commit()
    conn.close()
    _log_activity(user, "attendance_recorded", f"Attendance recorded for {payload.date} ({len(inserted_ids)} entries)", {
        "date": payload.date,
        "student_ids": inserted_ids,
    })

    return {"status": "ok", "message": "Attendance recorded", "count": len(payload.records)}

@app.post("/login")
def login(payload: LoginRequest):
    _ensure_passwords_for_students()
    passwords = _load_passwords()
    if passwords.get(payload.username) != payload.password:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    token = _create_session(payload.username)
    return {"status": "ok", "token": token}

@app.get("/auth/me")
def auth_me(request: Request):
    auth = request.headers.get("Authorization", "")
    if not auth.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Unauthorized")
    token = auth.replace("Bearer ", "").strip()
    session = _get_session(token)
    if not session:
        raise HTTPException(status_code=401, detail="Session expired")
    user = session["user"]
    user_key = str(user or "").lower()
    if user_key == "superuser" or user_key in STAFF_USERS:
        welcome = STAFF_USERS[user_key]["welcome"] if user_key in STAFF_USERS else ""
        return {"status": "ok", "user": user, "role": "superuser", "welcome_message": welcome}

    conn = get_connection()
    student = conn.execute(
        "SELECT student_name, course, batch FROM students WHERE student_id = ?",
        (user,),
    ).fetchone()
    conn.close()

    first_name = ""
    if student and student["student_name"]:
        first_name = str(student["student_name"]).strip().split(" ")[0]

    return {
        "status": "ok",
        "user": user,
        "role": "student",
        "student_name": student["student_name"] if student else "",
        "first_name": first_name,
        "course": student["course"] if student else "",
        "batch": student["batch"] if student else "",
    }


@app.post("/auth/change-password")
def auth_change_password(payload: ChangePasswordRequest, request: Request):
    user = _get_current_user(request)
    username = str((user or {}).get("user") or "").strip()
    if not username:
        raise HTTPException(status_code=401, detail="Unauthorized")
    _ensure_passwords_for_students()
    passwords = _load_passwords()
    current = str(payload.current_password or "")
    new_password = str(payload.new_password or "")
    if not current or not new_password:
        raise HTTPException(status_code=400, detail="current_password and new_password are required")
    if len(new_password) < 6:
        raise HTTPException(status_code=400, detail="New password must be at least 6 characters")
    if passwords.get(username) != current:
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    passwords[username] = new_password
    _save_passwords(passwords)
    _log_activity(user, "password_changed", f"Password changed for {username}", {"username": username})
    return {"status": "ok", "message": "Password updated"}


@app.post("/admin/users/password")
def admin_set_user_password(payload: AdminSetPasswordRequest, request: Request):
    user = _get_current_user(request)
    _require_superuser(user)
    username = str(payload.username or "").strip()
    new_password = str(payload.new_password or "")
    if not username or not new_password:
        raise HTTPException(status_code=400, detail="username and new_password are required")
    if len(new_password) < 6:
        raise HTTPException(status_code=400, detail="New password must be at least 6 characters")
    _ensure_passwords_for_students()
    passwords = _load_passwords()
    if username not in passwords:
        raise HTTPException(status_code=404, detail="User not found")
    passwords[username] = new_password
    _save_passwords(passwords)
    _log_activity(user, "password_reset", f"Password reset for {username}", {"username": username})
    return {"status": "ok", "message": "User password updated"}


@app.get("/activity/logs")
def activity_logs(request: Request):
    user = _get_current_user(request)
    _require_superuser(user)
    _ensure_activity_table()
    conn = get_connection()
    rows = conn.execute(
        """
        SELECT activity_id, action_type, description, payload_json, created_by, created_at, undone, undone_at
        FROM activity_log
        ORDER BY activity_id DESC
        LIMIT 200
        """
    ).fetchall()
    conn.close()
    out = []
    for r in rows:
        payload = {}
        try:
            payload = json.loads(r["payload_json"] or "{}")
        except Exception:
            payload = {}
        action_type = r["action_type"]
        undoable = (not bool(r["undone"])) and action_type in (
            "student_added",
            "students_batch_updated",
            "students_marked_alumni",
            "students_deleted",
            "admission_deleted",
            "attendance_recorded",
            "fee_recorded",
            "timetable_created",
            "interview_created",
            "announcement_created",
            "notification_created",
            "admission_submitted",
        )
        out.append({
            "activity_id": r["activity_id"],
            "action_type": action_type,
            "description": r["description"],
            "payload": payload,
            "created_by": r["created_by"],
            "created_at": r["created_at"],
            "undone": bool(r["undone"]),
            "undone_at": r["undone_at"],
            "undoable": undoable,
        })
    return out


@app.post("/activity/undo")
def activity_undo(payload: ActivityUndoRequest, request: Request):
    user = _get_current_user(request)
    _require_superuser(user)
    _ensure_activity_table()
    conn = get_connection()
    row = conn.execute(
        "SELECT activity_id, action_type, payload_json, undone FROM activity_log WHERE activity_id = ?",
        (payload.activity_id,),
    ).fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Activity not found")
    if int(row["undone"] or 0) == 1:
        conn.close()
        raise HTTPException(status_code=400, detail="Already undone")
    try:
        p = json.loads(row["payload_json"] or "{}")
    except Exception:
        p = {}
    action = row["action_type"]
    if action == "student_added":
        sid = str(p.get("student_id", "")).strip()
        if not sid:
            conn.close()
            raise HTTPException(status_code=400, detail="Undo payload missing student_id")
        conn.execute("DELETE FROM students WHERE student_id = ?", (sid,))
        passwords = _load_passwords()
        if sid in passwords:
            passwords.pop(sid, None)
            _save_passwords(passwords)
    elif action == "students_batch_updated":
        student_ids = p.get("student_ids", []) if isinstance(p.get("student_ids", []), list) else []
        previous_batches = p.get("previous_batches", {}) if isinstance(p.get("previous_batches", {}), dict) else {}
        if not student_ids:
            conn.close()
            raise HTTPException(status_code=400, detail="Undo payload missing student_ids")
        for sid in student_ids:
            prev = str(previous_batches.get(str(sid), "")).strip()
            if not prev:
                continue
            conn.execute("UPDATE students SET batch = ? WHERE student_id = ?", (prev, str(sid)))
            conn.execute("UPDATE attendance SET batch = ? WHERE student_id = ?", (prev, str(sid)))
    elif action == "students_marked_alumni":
        student_ids = p.get("student_ids", []) if isinstance(p.get("student_ids", []), list) else []
        previous_status = p.get("previous_status", {}) if isinstance(p.get("previous_status", {}), dict) else {}
        if not student_ids:
            conn.close()
            raise HTTPException(status_code=400, detail="Undo payload missing student_ids")
        for sid in student_ids:
            status = str(previous_status.get(str(sid), "Active")).strip() or "Active"
            conn.execute("UPDATE students SET status = ? WHERE student_id = ?", (status, str(sid)))
    elif action == "students_deleted":
        students = p.get("students", []) if isinstance(p.get("students", []), list) else []
        if not students:
            conn.close()
            raise HTTPException(status_code=400, detail="Undo payload missing students")
        for snapshot in students:
            student = snapshot.get("student", {}) if isinstance(snapshot, dict) else {}
            sid = str(student.get("student_id", "")).strip()
            if not sid:
                continue
            conn.execute(
                "INSERT OR IGNORE INTO students (student_id, student_name, course, batch, status) VALUES (?, ?, ?, ?, ?)",
                (
                    sid,
                    str(student.get("student_name", "")),
                    str(student.get("course", "")),
                    str(student.get("batch", "")),
                    str(student.get("status", "Active") or "Active"),
                ),
            )
            for att in snapshot.get("attendance", []) if isinstance(snapshot.get("attendance", []), list) else []:
                conn.execute(
                    """
                    INSERT OR IGNORE INTO attendance (student_id, student_name, course, batch, date, attendance_status, remarks)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        sid,
                        str(att.get("student_name", student.get("student_name", ""))),
                        str(att.get("course", student.get("course", ""))),
                        str(att.get("batch", student.get("batch", ""))),
                        str(att.get("date", "")),
                        str(att.get("attendance_status", "A")),
                        str(att.get("remarks", "")),
                    ),
                )
            for fee in snapshot.get("fees", []) if isinstance(snapshot.get("fees", []), list) else []:
                conn.execute(
                    """
                    INSERT INTO fees (student_id, amount_total, amount_paid, due_date, remarks, receipt_path)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (
                        sid,
                        float(fee.get("amount_total", 0) or 0),
                        float(fee.get("amount_paid", 0) or 0),
                        str(fee.get("due_date", "")),
                        str(fee.get("remarks", "")),
                        str(fee.get("receipt_path", "")),
                    ),
                )
        _ensure_passwords_for_students()
    elif action == "attendance_recorded":
        date = str(p.get("date", "")).strip()
        student_ids = p.get("student_ids", []) if isinstance(p.get("student_ids", []), list) else []
        if not date or not student_ids:
            conn.close()
            raise HTTPException(status_code=400, detail="Undo payload missing attendance details")
        for sid in student_ids:
            conn.execute("DELETE FROM attendance WHERE student_id = ? AND date = ?", (str(sid), date))
    elif action == "fee_recorded":
        fee_id = int(p.get("fee_id", 0) or 0)
        if not fee_id:
            conn.close()
            raise HTTPException(status_code=400, detail="Undo payload missing fee_id")
        conn.execute("DELETE FROM fees WHERE fee_id = ?", (fee_id,))
    elif action == "timetable_created":
        timetable_id = int(p.get("timetable_id", 0) or 0)
        if not timetable_id:
            conn.close()
            raise HTTPException(status_code=400, detail="Undo payload missing timetable_id")
        conn.execute("DELETE FROM timetable WHERE timetable_id = ?", (timetable_id,))
    elif action == "interview_created":
        interview_id = int(p.get("interview_id", 0) or 0)
        if not interview_id:
            conn.close()
            raise HTTPException(status_code=400, detail="Undo payload missing interview_id")
        conn.execute("DELETE FROM interview_stats WHERE interview_id = ?", (interview_id,))
    elif action == "announcement_created":
        announcement_id = int(p.get("announcement_id", 0) or 0)
        if not announcement_id:
            conn.close()
            raise HTTPException(status_code=400, detail="Undo payload missing announcement_id")
        conn.execute("DELETE FROM announcements WHERE announcement_id = ?", (announcement_id,))
    elif action == "notification_created":
        notification_id = int(p.get("notification_id", 0) or 0)
        if not notification_id:
            conn.close()
            raise HTTPException(status_code=400, detail="Undo payload missing notification_id")
        conn.execute("DELETE FROM notifications WHERE notification_id = ?", (notification_id,))
    elif action == "admission_submitted":
        admission_id = int(p.get("admission_id", 0) or 0)
        if not admission_id:
            conn.close()
            raise HTTPException(status_code=400, detail="Undo payload missing admission_id")
        _ensure_admissions_table()
        conn.execute("DELETE FROM admissions WHERE admission_id = ?", (admission_id,))
    elif action == "admission_deleted":
        admission = p.get("admission", {}) if isinstance(p.get("admission", {}), dict) else {}
        if not admission or not admission.get("admission_id"):
            conn.close()
            raise HTTPException(status_code=400, detail="Undo payload missing admission")
        _ensure_admissions_table()
        conn.execute(
            """
            INSERT OR REPLACE INTO admissions (
                admission_id, full_name, first_name, middle_name, last_name, phone, email, blood_group, age, dob,
                aadhaar_number, nationality, father_name, father_phone, father_occupation, father_email,
                mother_name, mother_phone, mother_occupation, mother_email, correspondence_address,
                permanent_address, course, academic_details_json, admission_pdf_path, admission_pdf_bytes, status, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                int(admission.get("admission_id", 0) or 0),
                str(admission.get("full_name", "")),
                str(admission.get("first_name", "")),
                str(admission.get("middle_name", "")),
                str(admission.get("last_name", "")),
                str(admission.get("phone", "")),
                str(admission.get("email", "")),
                str(admission.get("blood_group", "")),
                int(admission.get("age", 0) or 0),
                str(admission.get("dob", "")),
                str(admission.get("aadhaar_number", "")),
                str(admission.get("nationality", "")),
                str(admission.get("father_name", "")),
                str(admission.get("father_phone", "")),
                str(admission.get("father_occupation", "")),
                str(admission.get("father_email", "")),
                str(admission.get("mother_name", "")),
                str(admission.get("mother_phone", "")),
                str(admission.get("mother_occupation", "")),
                str(admission.get("mother_email", "")),
                str(admission.get("correspondence_address", "")),
                str(admission.get("permanent_address", "")),
                str(admission.get("course", "")),
                str(admission.get("academic_details_json", "[]")),
                str(admission.get("admission_pdf_path", "")),
                int(admission.get("admission_pdf_bytes", 0) or 0),
                str(admission.get("status", "new")),
                str(admission.get("created_at", "")),
            ),
        )
    else:
        conn.close()
        raise HTTPException(status_code=400, detail="This action is not undoable")
    conn.execute("UPDATE activity_log SET undone = 1, undone_at = datetime('now') WHERE activity_id = ?", (payload.activity_id,))
    conn.commit()
    conn.close()
    _log_activity(user, "undo_action", f"Undid activity #{payload.activity_id}", {"target_activity_id": payload.activity_id})
    return {"status": "ok", "message": "Undo successful"}

@app.post("/attendance/sync")
def sync_attendance_from_excel(request: Request):
    user = _get_current_user(request)
    _require_superuser(user)
    try:
        result = import_attendance(reset_attendance=True)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    _log_activity(
        user,
        "attendance_synced_excel",
        f"Attendance synced from Excel ({result['inserted']} inserted, {result['skipped']} skipped)",
        {"inserted": result["inserted"], "skipped": result["skipped"]},
    )
    return {"status": "ok", "message": "Attendance synced from Excel", "inserted": result["inserted"], "skipped": result["skipped"]}


@app.post("/attendance/sync/upload")
async def sync_attendance_upload(file: UploadFile = File(...), request: Request = None):
    user = _get_current_user(request)
    _require_superuser(user)
    filename = file.filename or ""
    if not filename.lower().endswith(".csv"):
        return {
            "status": "uploaded_only",
            "message": "Please upload CSV for automatic parsing.",
            "supported_parse_format": "csv",
        }

    raw = (await file.read()).decode("utf-8-sig", errors="ignore")
    reader = csv.DictReader(raw.splitlines())
    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="CSV has no headers")
    headers = [h.strip().lower() for h in reader.fieldnames if h]
    required = ["student_id", "student_name", "course", "batch", "date", "attendance_status"]
    missing = [c for c in required if c not in headers]
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing CSV columns: {', '.join(missing)}")

    conn = get_connection()
    conn.execute("DELETE FROM attendance")
    inserted = 0
    skipped = 0
    for row in reader:
        row_norm = {str(k).strip().lower(): v for k, v in row.items()}
        student_id = str(row_norm.get("student_id", "")).strip()
        student_name = str(row_norm.get("student_name", "")).strip()
        course = str(row_norm.get("course", "")).strip()
        batch = str(row_norm.get("batch", "")).strip()
        date = normalize_date_to_iso(row_norm.get("date", ""))
        status = normalize_attendance_status(row_norm.get("attendance_status", ""))
        remarks = str(row_norm.get("remarks", "")).strip()
        if not student_id or not date:
            skipped += 1
            continue
        conn.execute(
            "INSERT OR IGNORE INTO students (student_id, student_name, course, batch) VALUES (?, ?, ?, ?)",
            (student_id, student_name, course, batch),
        )
        cur = conn.execute(
            """
            INSERT OR IGNORE INTO attendance
            (student_id, student_name, course, batch, date, attendance_status, remarks)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (student_id, student_name, course, batch, date, status, remarks),
        )
        if cur.rowcount == 1:
            inserted += 1
        else:
            skipped += 1
    conn.commit()
    conn.close()
    _log_activity(
        user,
        "attendance_synced_csv",
        f"Attendance synced from CSV ({inserted} inserted, {skipped} skipped)",
        {"inserted": inserted, "skipped": skipped, "filename": filename},
    )
    return {"status": "ok", "message": "Attendance synced from uploaded CSV", "inserted": inserted, "skipped": skipped}

@app.post("/fees/record")
async def record_fee(
    student_id: str = Form(...),
    amount_paid: float = Form(...),
    amount_total: float = Form(None),
    due_date: str = Form(None),
    remarks: str = Form(""),
    payment_mode: str = Form(""),
    bank_name: str = Form(""),
    txn_utr_no: str = Form(""),
    bank_ref_no: str = Form(""),
    transaction_type: str = Form(""),
    receipt: UploadFile = File(None),
    request: Request = None,
):
    user = _get_current_user(request)
    _require_superuser(user)
    conn = get_connection()
    _ensure_fee_receipt_columns(conn)
    info = _student_financials(conn, student_id)
    if not info:
        conn.close()
        raise HTTPException(status_code=404, detail="Student not found")
    if amount_total is None:
        amount_total = float(info["total"])
    mode = (payment_mode or "").strip().upper()
    if not mode:
        mode = "OFFLINE"
    bank_name = (bank_name or "").strip()
    txn_utr_no = (txn_utr_no or "").strip()
    bank_ref_no = (bank_ref_no or "").strip()
    txn_type = (transaction_type or "").strip().upper()
    if not txn_type:
        txn_type = mode
    if not bank_name:
        bank_name = "Cash" if mode == "CASH" else ("Razorpay" if mode == "ONLINE" else "NA")
    if not txn_utr_no:
        txn_utr_no = "NA"
    if not bank_ref_no:
        bank_ref_no = "NA"

    receipt_path = None
    if receipt is not None:
        RECEIPTS_DIR.mkdir(parents=True, exist_ok=True)
        ext = Path(receipt.filename).suffix
        filename = f"{student_id}_{uuid4().hex}{ext}"
        save_path = RECEIPTS_DIR / filename
        content = await receipt.read()
        save_path.write_bytes(content)
        receipt_path = str(save_path)

    conn = get_connection()
    cur = conn.execute(
        """
        INSERT INTO fees (
            student_id,
            amount_total,
            amount_paid,
            due_date,
            remarks,
            receipt_path,
            payment_mode,
            bank_name,
            txn_utr_no,
            bank_ref_no,
            transaction_type
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            student_id,
            amount_total,
            amount_paid,
            due_date,
            remarks,
            receipt_path,
            mode,
            bank_name,
            txn_utr_no,
            bank_ref_no,
            txn_type,
        ),
    )
    fee_id = cur.lastrowid
    conn.commit()
    conn.close()
    _log_activity(user, "fee_recorded", f"Fee recorded for {student_id} (INR {amount_paid})", {
        "fee_id": fee_id,
        "student_id": str(student_id),
        "amount_paid": amount_paid,
    })

    return {"status": "ok", "message": "Fee recorded", "fee_id": fee_id}

@app.get("/fees/recent")
def recent_fees(request: Request):
    user = _get_current_user(request)
    conn = get_connection()
    if user and not _is_superuser(user):
        rows = conn.execute(
            """
            SELECT fee_id, student_id, amount_total, amount_paid, due_date, remarks
            FROM fees
            WHERE student_id = ?
            ORDER BY fee_id DESC
            LIMIT 20
            """,
            (user["user"],),
        ).fetchall()
    else:
        rows = conn.execute(
            """
            SELECT fee_id, student_id, amount_total, amount_paid, due_date, remarks
            FROM fees
            ORDER BY fee_id DESC
            LIMIT 20
            """
        ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.get("/fees/admin/policies")
def fees_admin_policies(request: Request):
    user = _get_current_user(request)
    _require_superuser(user)
    _ensure_fee_policies_table()
    conn = get_connection()
    rows = conn.execute(
        "SELECT student_id, concession_amount, due_date, updated_at FROM fee_policies ORDER BY student_id ASC"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.post("/fees/admin/policy")
def fees_admin_policy(payload: FeePolicyUpdateRequest, request: Request):
    user = _get_current_user(request)
    _require_superuser(user)
    student_id = str(payload.student_id or "").strip()
    if not student_id:
        raise HTTPException(status_code=400, detail="student_id is required")
    _ensure_fee_policies_table()
    conn = get_connection()
    info_before = _student_financials(conn, student_id)
    if not info_before:
        conn.close()
        raise HTTPException(status_code=404, detail="Student not found")

    concession = max(float(payload.concession_amount or 0), 0.0)
    concession = min(concession, max(float(info_before["base_total"]), 0.0))
    due_date = (payload.due_date or "").strip()
    if due_date and not re.match(r"^\d{4}-\d{2}-\d{2}$", due_date):
        conn.close()
        raise HTTPException(status_code=400, detail="due_date must be YYYY-MM-DD")
    due_date = due_date or None

    conn.execute(
        """
        INSERT INTO fee_policies (student_id, concession_amount, due_date, updated_at)
        VALUES (?, ?, ?, datetime('now'))
        ON CONFLICT(student_id) DO UPDATE SET
          concession_amount = excluded.concession_amount,
          due_date = excluded.due_date,
          updated_at = datetime('now')
        """,
        (student_id, concession, due_date),
    )
    info_after = _student_financials(conn, student_id)
    conn.commit()
    conn.close()
    _log_activity(user, "fee_policy_updated", f"Updated fee policy for {student_id}", {
        "student_id": student_id,
        "concession_amount": concession,
        "due_date": due_date,
    })
    return {
        "status": "ok",
        "student_id": student_id,
        "concession_amount": info_after["concession_amount"],
        "due_date": info_after["due_date"],
        "total": info_after["total"],
        "due": info_after["due"],
    }


@app.post("/fees/reminders")
def send_fee_reminders(payload: FeeRemindersRequest, request: Request):
    user = _get_current_user(request)
    _require_superuser(user)
    conn = get_connection()
    students = conn.execute("SELECT student_id FROM students").fetchall()

    from datetime import datetime, timedelta
    today = datetime.now()
    max_date = today + timedelta(days=payload.days)
    today_str = today.strftime("%Y-%m-%d")
    max_date_str = max_date.strftime("%Y-%m-%d")

    sent = 0
    for s in students:
        info = _student_financials(conn, s["student_id"])
        if not info or info["due"] <= 0 or not info["due_date"]:
            continue
        if today_str <= info["due_date"] <= max_date_str:
            title = "Fee Reminder"
            message = f"Your fee balance is INR {info['due']}. Due date: {info['due_date']}."
            conn.execute(
                "INSERT INTO notifications (title, message, level, target_user) VALUES (?, ?, ?, ?)",
                (title, message, "info", s["student_id"]),
            )
            sent += 1
    conn.commit()
    conn.close()
    _log_activity(user, "fee_reminders_sent", f"Sent {sent} fee reminders", {"days": payload.days, "sent": sent})
    return {"status": "ok", "sent": sent}


@app.post("/fees/admin/reset-unpaid")
def fees_admin_reset_unpaid(request: Request):
    user = _get_current_user(request)
    _require_superuser(user)
    _ensure_fee_policies_table()
    conn = get_connection()
    students = conn.execute("SELECT student_id, course FROM students ORDER BY student_id ASC").fetchall()
    existing = conn.execute(
        "SELECT student_id, COALESCE(MAX(amount_total),0) AS max_total FROM fees GROUP BY student_id"
    ).fetchall()
    existing_by_sid = {str(r["student_id"]): float(r["max_total"] or 0) for r in existing}
    policies = conn.execute("SELECT student_id, concession_amount, due_date FROM fee_policies").fetchall()
    policy_by_sid = {str(r["student_id"]): dict(r) for r in policies}

    conn.execute("DELETE FROM fees")

    inserted = 0
    for row in students:
        sid = str(row["student_id"])
        planned = _course_fee_inr(row["course"])
        base_total = float(planned if planned is not None else existing_by_sid.get(sid, 0.0))
        policy = policy_by_sid.get(sid) or {}
        concession = min(max(float(policy.get("concession_amount", 0) or 0), 0.0), max(base_total, 0.0))
        effective_total = max(base_total - concession, 0.0)
        due_date = policy.get("due_date")
        conn.execute(
            """
            INSERT INTO fees (student_id, amount_total, amount_paid, due_date, remarks)
            VALUES (?, ?, 0, ?, ?)
            """,
            (sid, effective_total, due_date, "Fee reset to unpaid by staff"),
        )
        inserted += 1
    conn.commit()
    conn.close()
    _log_activity(user, "fees_reset_unpaid", f"Reset fees to 100% unpaid for {inserted} students", {
        "students_count": inserted,
    })
    return {"status": "ok", "message": f"Reset to unpaid for {inserted} students", "students_count": inserted}


@app.get("/fees/{fee_id}/invoice")
def fee_invoice(fee_id: int, request: Request):
    user = _get_current_user(request)
    conn = get_connection()
    _ensure_fee_receipt_columns(conn)
    row = conn.execute(
        """
        SELECT
          f.fee_id,
          f.student_id,
          f.amount_total,
          f.amount_paid,
          f.remarks,
          f.payment_mode,
          f.bank_name,
          f.txn_utr_no,
          f.bank_ref_no,
          f.transaction_type,
          f.created_at,
          s.student_name,
          s.course
        FROM fees f
        LEFT JOIN students s ON s.student_id = f.student_id
        WHERE f.fee_id = ?
        """,
        (fee_id,),
    ).fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Fee entry not found")

    _require_self_or_superuser(user, row["student_id"])
    info = _student_financials(conn, row["student_id"])
    conn.close()
    refs = _parse_razorpay_refs(row["remarks"])
    default_due = max(float(row["amount_total"] or 0) - float(row["amount_paid"] or 0), 0.0)

    return {
        "invoice": {
            "invoice_no": f"AAI-INV-{int(row['fee_id'])}",
            "date": _to_iso_date(row["created_at"]),
            "student_id": row["student_id"] or "",
            "student_name": row["student_name"] or "",
            "course": row["course"] or "",
            "payment_id": refs["payment_id"],
            "order_id": refs["order_id"],
            "payment_mode": row["payment_mode"] or "",
            "bank_name": row["bank_name"] or "",
            "txn_utr_no": row["txn_utr_no"] or "",
            "bank_ref_no": row["bank_ref_no"] or "",
            "transaction_type": row["transaction_type"] or "",
            "amount_paid": float(row["amount_paid"] or 0),
            "amount_total": float(row["amount_total"] or 0),
            "balance_due": float(info["due"]) if info else default_due,
            "concession_amount": float(info["concession_amount"]) if info else 0.0,
        }
    }


@app.get("/fees/summary")
def fees_summary(request: Request):
    user = _get_current_user(request)
    conn = get_connection()
    if user and not _is_superuser(user):
        info = _student_financials(conn, user["user"])
        conn.close()
        if not info:
            return {"total": 0, "paid": 0, "due": 0, "transactions": 0}
        return {
            "total": info["total"],
            "paid": info["paid"],
            "due": info["due"],
            "transactions": info["transactions"],
            "course": info["student"]["course"],
            "concession_amount": info["concession_amount"],
            "due_date": info["due_date"],
            "gst_percent": 18,
        }
    students = conn.execute("SELECT student_id FROM students").fetchall()
    total = 0.0
    paid = 0.0
    for s in students:
        info = _student_financials(conn, s["student_id"])
        if not info:
            continue
        total += float(info["total"])
        paid += float(info["paid"])
    row = conn.execute("SELECT COUNT(*) AS transactions FROM fees").fetchone()
    conn.close()
    return {
        "total": total,
        "paid": paid,
        "due": total - paid,
        "transactions": int(row["transactions"]),
    }


@app.get("/payments/gateway-status")
def payment_gateway_status(request: Request):
    _ = _get_current_user(request)
    return {
        "enabled": bool(RAZORPAY_KEY_ID and RAZORPAY_KEY_SECRET),
        "provider": "razorpay",
        "message": "Razorpay ready" if (RAZORPAY_KEY_ID and RAZORPAY_KEY_SECRET) else "Razorpay keys not configured.",
        "key_id": RAZORPAY_KEY_ID if RAZORPAY_KEY_ID else None,
    }


@app.post("/payments/razorpay/order")
def create_razorpay_order(payload: RazorpayOrderRequest, request: Request):
    user = _get_current_user(request)
    if not (RAZORPAY_KEY_ID and RAZORPAY_KEY_SECRET):
        raise HTTPException(status_code=503, detail="Razorpay is not configured")

    student_id = payload.student_id or user["user"]
    _require_self_or_superuser(user, student_id)

    conn = get_connection()
    info = _student_financials(conn, student_id)
    conn.close()
    if not info:
        raise HTTPException(status_code=404, detail="Student not found")

    due = float(info["due"])
    if due <= 0:
        raise HTTPException(status_code=400, detail="No due amount")

    amount_inr = float(payload.amount_inr) if payload.amount_inr is not None else due
    amount_inr = max(1.0, min(amount_inr, due))
    amount_paise = int(round(amount_inr * 100))
    receipt = f"fee-{student_id}-{int(time.time())}"

    try:
        order = _create_razorpay_order(
            amount_paise=amount_paise,
            receipt=receipt,
            notes={"student_id": student_id, "course": info["student"]["course"] or ""},
        )
    except urlerror.HTTPError as e:
        detail = e.read().decode("utf-8", errors="ignore")
        raise HTTPException(status_code=502, detail=f"Razorpay error: {detail}")
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Failed to create Razorpay order: {e}")

    return {
        "key_id": RAZORPAY_KEY_ID,
        "order": order,
        "student_id": student_id,
        "amount_inr": amount_inr,
        "due_inr": due,
        "student_name": info["student"]["student_name"],
    }


@app.post("/payments/razorpay/verify")
def verify_razorpay_payment(payload: RazorpayVerifyRequest, request: Request):
    user = _get_current_user(request)
    _require_self_or_superuser(user, payload.student_id)
    if not (RAZORPAY_KEY_ID and RAZORPAY_KEY_SECRET):
        raise HTTPException(status_code=503, detail="Razorpay is not configured")

    signing_data = f"{payload.razorpay_order_id}|{payload.razorpay_payment_id}".encode("utf-8")
    generated_signature = hmac.new(
        RAZORPAY_KEY_SECRET.encode("utf-8"),
        signing_data,
        hashlib.sha256,
    ).hexdigest()

    if not hmac.compare_digest(generated_signature, payload.razorpay_signature):
        raise HTTPException(status_code=400, detail="Invalid payment signature")

    conn = get_connection()
    _ensure_fee_receipt_columns(conn)
    info = _student_financials(conn, payload.student_id)
    if not info:
        conn.close()
        raise HTTPException(status_code=404, detail="Student not found")

    amount_paid = max(0.0, float(payload.amount_paid_inr))
    amount_paid = min(amount_paid, float(info["due"]))
    remarks = (
        f"Razorpay payment_id={payload.razorpay_payment_id}, "
        f"order_id={payload.razorpay_order_id}"
    )
    cur = conn.execute(
        """
        INSERT INTO fees (
            student_id,
            amount_total,
            amount_paid,
            remarks,
            payment_mode,
            bank_name,
            txn_utr_no,
            bank_ref_no,
            transaction_type
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            payload.student_id,
            float(info["total"]),
            amount_paid,
            remarks,
            "ONLINE",
            "Razorpay",
            payload.razorpay_payment_id,
            payload.razorpay_order_id,
            "ONLINE",
        ),
    )
    fee_id = cur.lastrowid
    balance_due = max(float(info["due"]) - amount_paid, 0.0)
    conn.commit()
    conn.close()
    return {
        "status": "ok",
        "message": "Payment verified and recorded",
        "amount_paid_inr": amount_paid,
        "invoice": {
            "invoice_no": f"AAI-INV-{int(fee_id)}",
            "date": time.strftime("%Y-%m-%d"),
            "student_id": payload.student_id,
            "student_name": info["student"]["student_name"],
            "course": info["student"]["course"] or "",
            "payment_id": payload.razorpay_payment_id,
            "order_id": payload.razorpay_order_id,
            "payment_mode": "ONLINE",
            "bank_name": "Razorpay",
            "txn_utr_no": payload.razorpay_payment_id,
            "bank_ref_no": payload.razorpay_order_id,
            "transaction_type": "ONLINE",
            "amount_paid": amount_paid,
            "amount_total": float(info["total"]),
            "balance_due": balance_due,
            "concession_amount": float(info["concession_amount"]),
        },
    }


@app.get("/timetable")
def list_timetable(request: Request, course: str = "", batch: str = ""):
    user = _get_current_user(request)
    conn = get_connection()
    params = []
    where = []

    if user and not _is_superuser(user):
        student = conn.execute(
            "SELECT course, batch FROM students WHERE student_id = ?",
            (user["user"],),
        ).fetchone()
        if student:
            course = student["course"] or ""
            batch = student["batch"] or ""

    if course:
        where.append("(course = ? OR course = '')")
        params.append(course)
    if batch:
        where.append("(batch = ? OR batch = '')")
        params.append(batch)

    query = "SELECT * FROM timetable"
    if where:
        query += " WHERE " + " AND ".join(where)
    query += " ORDER BY day_of_week ASC, start_time ASC, timetable_id DESC"
    rows = conn.execute(query, tuple(params)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.post("/timetable")
def create_timetable(entry: TimetableEntryRequest, request: Request):
    user = _get_current_user(request)
    _require_superuser(user)
    cur = conn.execute(
        """
        INSERT INTO timetable
        (title, day_of_week, start_time, end_time, course, batch, location, instructor)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            entry.title,
            entry.day_of_week,
            entry.start_time,
            entry.end_time,
            entry.course or "",
            entry.batch or "",
            entry.location or "",
            entry.instructor or "",
        ),
    )
    timetable_id = cur.lastrowid
    conn.commit()
    conn.close()
    _log_activity(
        user,
        "timetable_created",
        f"Timetable entry added: {entry.title} ({entry.day_of_week})",
        {"timetable_id": timetable_id, "title": entry.title, "day_of_week": entry.day_of_week},
    )
    return {"status": "ok", "message": "Timetable entry created"}


@app.get("/interviews")
def list_interviews(request: Request):
    user = _get_current_user(request)
    conn = get_connection()
    manual_rows = conn.execute(
        """
        SELECT * FROM interview_stats
        ORDER BY interview_date DESC, interview_id DESC
        """
    ).fetchall()

    if user and not _is_superuser(user):
        attendance_rows = conn.execute(
            """
            SELECT attendance_id, student_id, student_name, date, remarks
            FROM attendance
            WHERE student_id = ?
              AND remarks IS NOT NULL
              AND TRIM(remarks) <> ''
              AND lower(remarks) LIKE '%interview%'
            ORDER BY date DESC, attendance_id DESC
            """,
            (user["user"],),
        ).fetchall()
    else:
        attendance_rows = conn.execute(
            """
            SELECT attendance_id, student_id, student_name, date, remarks
            FROM attendance
            WHERE remarks IS NOT NULL
              AND TRIM(remarks) <> ''
              AND lower(remarks) LIKE '%interview%'
            ORDER BY date DESC, attendance_id DESC
            """
        ).fetchall()
    conn.close()

    interviews = []
    for r in manual_rows:
        item = dict(r)
        item["source"] = "manual"
        interviews.append(item)

    for r in attendance_rows:
        remark = r["remarks"] or ""
        interviews.append(
            {
                "interview_id": f"attendance-{r['attendance_id']}",
                "airline_name": _extract_airline_from_interview_remark(remark),
                "interview_date": r["date"],
                "notes": remark,
                "source": "attendance_remark",
                "student_id": r["student_id"],
                "student_name": r["student_name"],
            }
        )

    interviews.sort(key=lambda x: str(x.get("interview_date", "")), reverse=True)
    return interviews


@app.post("/interviews")
def create_interview(entry: InterviewStatRequest, request: Request):
    user = _get_current_user(request)
    _require_superuser(user)
    conn = get_connection()
    cur = conn.execute(
        """
        INSERT INTO interview_stats (airline_name, interview_date, notes)
        VALUES (?, ?, ?)
        """,
        (entry.airline_name, entry.interview_date, entry.notes or ""),
    )
    interview_id = cur.lastrowid
    conn.commit()
    conn.close()
    _log_activity(
        user,
        "interview_created",
        f"Interview record added: {entry.airline_name} ({entry.interview_date})",
        {"interview_id": interview_id, "airline_name": entry.airline_name, "interview_date": entry.interview_date},
    )
    return {"status": "ok", "message": "Interview stat created"}


@app.get("/announcements")
def list_announcements(request: Request, limit: int = 20):
    _ = _get_current_user(request)
    conn = get_connection()
    rows = conn.execute(
        """
        SELECT * FROM announcements
        ORDER BY announcement_id DESC
        LIMIT ?
        """,
        (max(1, min(limit, 100)),),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.post("/announcements")
def create_announcement(payload: AnnouncementRequest, request: Request):
    user = _get_current_user(request)
    _require_superuser(user)
    conn = get_connection()
    cur = conn.execute(
        """
        INSERT INTO announcements (title, message, created_by)
        VALUES (?, ?, ?)
        """,
        (payload.title, payload.message, user["user"]),
    )
    announcement_id = cur.lastrowid
    conn.commit()
    conn.close()
    _log_activity(
        user,
        "announcement_created",
        f"Announcement posted: {payload.title}",
        {"announcement_id": announcement_id, "title": payload.title},
    )
    return {"status": "ok", "message": "Announcement created"}


@app.get("/notifications")
def list_notifications(request: Request, limit: int = 30):
    user = _get_current_user(request)
    conn = get_connection()
    uid = user["user"]
    rows = conn.execute(
        """
        SELECT
          n.notification_id, n.title, n.message, n.level, n.target_user, n.created_at,
          CASE WHEN nr.user_id IS NULL THEN 0 ELSE 1 END AS is_read
        FROM notifications n
        LEFT JOIN notification_reads nr
          ON nr.notification_id = n.notification_id AND nr.user_id = ?
        WHERE n.target_user IS NULL OR n.target_user = ?
        ORDER BY n.notification_id DESC
        LIMIT ?
        """,
        (uid, uid, max(1, min(limit, 100))),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.post("/notifications")
def create_notification(payload: NotificationRequest, request: Request):
    user = _get_current_user(request)
    _require_superuser(user)
    conn = get_connection()
    cur = conn.execute(
        """
        INSERT INTO notifications (title, message, level, target_user)
        VALUES (?, ?, ?, ?)
        """,
        (
            payload.title,
            payload.message,
            payload.level or "info",
            payload.target_user,
        ),
    )
    notification_id = cur.lastrowid
    conn.commit()
    conn.close()
    _log_activity(
        user,
        "notification_created",
        f"Notification created: {payload.title}",
        {"notification_id": notification_id, "title": payload.title, "target_user": payload.target_user},
    )
    return {"status": "ok", "message": "Notification created"}


@app.get("/tests")
def list_tests(request: Request):
    user = _get_current_user(request)
    _ensure_tests_tables()
    conn = get_connection()
    if _is_superuser(user):
        rows = conn.execute(
            """
            SELECT
                t.test_id,
                t.title,
                t.duration_minutes,
                t.created_at,
                (SELECT COUNT(*) FROM test_questions q WHERE q.test_id = t.test_id) AS question_count,
                (SELECT COUNT(*) FROM test_assignments a WHERE a.test_id = t.test_id) AS assignment_count,
                (SELECT COUNT(*) FROM test_attempts x WHERE x.test_id = t.test_id) AS attempt_count,
                (SELECT COUNT(*) FROM test_attempts x WHERE x.test_id = t.test_id AND x.malpractice_flag = 1) AS malpractice_count
            FROM tests t
            WHERE t.is_active = 1
            ORDER BY t.test_id DESC
            """
        ).fetchall()
    else:
        uid = user["user"]
        rows = conn.execute(
            """
            SELECT
                t.test_id,
                t.title,
                t.duration_minutes,
                (
                    SELECT status FROM test_attempts x
                    WHERE x.test_id = t.test_id AND x.student_id = ?
                    ORDER BY x.attempt_id DESC LIMIT 1
                ) AS attempt_status
            FROM tests t
            WHERE t.is_active = 1
                AND (
                    NOT EXISTS (SELECT 1 FROM test_assignments a WHERE a.test_id = t.test_id)
                    OR EXISTS (SELECT 1 FROM test_assignments a WHERE a.test_id = t.test_id AND a.student_id = ?)
                )
            ORDER BY t.test_id DESC
            """,
            (uid, uid),
        ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.get("/tests/{test_id}")
def test_detail(test_id: int, request: Request):
    user = _get_current_user(request)
    _ensure_tests_tables()
    conn = get_connection()
    row = conn.execute(
        "SELECT test_id, title, description, duration_minutes FROM tests WHERE test_id = ? AND is_active = 1",
        (test_id,),
    ).fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Test not found")

    if not _is_superuser(user):
        uid = user["user"]
        assigned = conn.execute(
            """
            SELECT 1 AS ok WHERE
            NOT EXISTS (SELECT 1 FROM test_assignments a WHERE a.test_id = ?)
            OR EXISTS (SELECT 1 FROM test_assignments a WHERE a.test_id = ? AND a.student_id = ?)
            """,
            (test_id, test_id, uid),
        ).fetchone()
        if not assigned:
            conn.close()
            raise HTTPException(status_code=403, detail="Forbidden")

    q_rows = conn.execute(
        "SELECT question_id, question_order, question_text, option_a, option_b, option_c, option_d, correct_answer FROM test_questions WHERE test_id = ? ORDER BY question_order ASC",
        (test_id,),
    ).fetchall()
    conn.close()

    questions = []
    for q in q_rows:
        item = dict(q)
        if not _is_superuser(user):
            item.pop("correct_answer", None)
        questions.append(item)

    res = dict(row)
    res["questions"] = questions
    return res


@app.get("/tests/{test_id}/attempts")
def test_attempts_by_test(test_id: int, request: Request):
    user = _get_current_user(request)
    _require_superuser(user)
    _ensure_tests_tables()
    conn = get_connection()
    rows = conn.execute(
        """
        SELECT
            attempt_id, test_id, student_id, start_time, submitted_at, status, score, total_points, malpractice_count, malpractice_flag
        FROM test_attempts
        WHERE test_id = ?
        ORDER BY attempt_id DESC
        """,
        (test_id,),
    ).fetchall()

    out = []
    for r in rows:
        item = dict(r)
        events = conn.execute(
            "SELECT event_id, event_type, details, created_at FROM test_malpractice_events WHERE attempt_id = ? ORDER BY event_id ASC",
            (r["attempt_id"],),
        ).fetchall()
        item["malpractice_events"] = [dict(e) for e in events]
        out.append(item)
    conn.close()
    return out


@app.post("/tests/{test_id}/start")
def test_start(test_id: int, request: Request):
    user = _get_current_user(request)
    if _is_superuser(user):
        raise HTTPException(status_code=403, detail="Staff cannot take tests")
    _ensure_tests_tables()
    conn = get_connection()
    test = conn.execute("SELECT test_id, title, duration_minutes FROM tests WHERE test_id = ? AND is_active = 1", (test_id,)).fetchone()
    if not test:
        conn.close()
        raise HTTPException(status_code=404, detail="Test not found")

    uid = user["user"]
    assigned = conn.execute(
        """
        SELECT 1 AS ok WHERE
        NOT EXISTS (SELECT 1 FROM test_assignments a WHERE a.test_id = ?)
        OR EXISTS (SELECT 1 FROM test_assignments a WHERE a.test_id = ? AND a.student_id = ?)
        """,
        (test_id, test_id, uid),
    ).fetchone()
    if not assigned:
        conn.close()
        raise HTTPException(status_code=403, detail="Forbidden")

    attempt = conn.execute(
        "SELECT * FROM test_attempts WHERE test_id = ? AND student_id = ? AND status = 'in_progress' ORDER BY attempt_id DESC LIMIT 1",
        (test_id, uid),
    ).fetchone()

    if not attempt:
        cur = conn.execute("INSERT INTO test_attempts (test_id, student_id, status) VALUES (?, ?, 'in_progress')", (test_id, uid))
        attempt_id = cur.lastrowid
        attempt = conn.execute("SELECT * FROM test_attempts WHERE attempt_id = ?", (attempt_id,)).fetchone()
        _log_activity(user, "test_started", f"Test attempt started (test #{test_id})", {"test_id": test_id, "attempt_id": attempt_id})

    q_rows = conn.execute("SELECT question_id, question_order, question_text, option_a, option_b, option_c, option_d FROM test_questions WHERE test_id = ? ORDER BY question_order ASC", (test_id,)).fetchall()

    # Simple order for now, modular logic handles shuffle if needed
    questions = [dict(q) for q in q_rows]

    ans_rows = conn.execute("SELECT question_id, answer_text FROM test_attempt_answers WHERE attempt_id = ?", (attempt["attempt_id"],)).fetchall()
    answers = {r["question_id"]: r["answer_text"] for r in ans_rows}

    from datetime import datetime
    start_dt = datetime.fromisoformat(attempt["start_time"].replace(" ", "T"))
    import time
    start_epoch = int(start_dt.timestamp())
    ends_at_epoch = start_epoch + int(test["duration_minutes"]) * 60

    conn.close()
    return {
        "attempt_id": attempt["attempt_id"],
        "test_id": test_id,
        "title": test["title"],
        "status": attempt["status"],
        "ends_at_epoch": ends_at_epoch,
        "answers": answers,
        "questions": questions
    }


@app.post("/tests/attempts/{attempt_id}/submit")
def test_submit(attempt_id: int, payload: TestSubmissionRequest, request: Request):
    user = _get_current_user(request)
    if _is_superuser(user):
        raise HTTPException(status_code=403, detail="Staff cannot submit tests")
    _ensure_tests_tables()
    conn = get_connection()
    attempt = conn.execute("SELECT * FROM test_attempts WHERE attempt_id = ?", (attempt_id,)).fetchone()
    if not attempt:
        conn.close()
        raise HTTPException(status_code=404, detail="Attempt not found")
    if attempt["student_id"] != user["user"]:
        conn.close()
        raise HTTPException(status_code=403, detail="Forbidden")
    if attempt["status"] == "submitted":
        conn.close()
        raise HTTPException(status_code=400, detail="Already submitted")

    conn.execute("DELETE FROM test_attempt_answers WHERE attempt_id = ?", (attempt_id,))
    q_rows = conn.execute("SELECT question_id, correct_answer, points FROM test_questions WHERE test_id = ?", (attempt["test_id"],)).fetchall()
    by_id = {r["question_id"]: r for r in q_rows}

    score = 0
    total = sum(r["points"] for r in q_rows)
    for item in payload.answers:
        qid = item.get("question_id")
        if qid not in by_id: continue
        given = str(item.get("answer", "")).strip().upper()
        correct = str(by_id[qid]["correct_answer"]).strip().upper()
        ok = (given == correct)
        points = by_id[qid]["points"] if ok else 0
        score += points
        conn.execute("INSERT INTO test_attempt_answers (attempt_id, question_id, answer_text, is_correct, points_awarded) VALUES (?, ?, ?, ?, ?)",
                     (attempt_id, qid, given, 1 if ok else 0, points))

    conn.execute("UPDATE test_attempts SET status = 'submitted', submitted_at = datetime('now'), score = ?, total_points = ? WHERE attempt_id = ?",
                 (score, total, attempt_id))
    conn.commit()
    conn.close()
    _log_activity(user, "test_submitted", f"Test attempt submitted (attempt #{attempt_id})", {"attempt_id": attempt_id, "score": score, "total": total})
    return {"status": "ok", "score": score, "total_points": total}


@app.post("/tests/attempts/{attempt_id}/malpractice")
def test_malpractice(attempt_id: int, payload: MalpracticeRequest, request: Request):
    user = _get_current_user(request)
    _ensure_tests_tables()
    conn = get_connection()
    attempt = conn.execute("SELECT * FROM test_attempts WHERE attempt_id = ?", (attempt_id,)).fetchone()
    if not attempt:
        conn.close()
        raise HTTPException(status_code=404, detail="Attempt not found")
    if attempt["student_id"] != user["user"]:
        conn.close()
        raise HTTPException(status_code=403, detail="Forbidden")

    conn.execute("INSERT INTO test_malpractice_events (attempt_id, event_type, details) VALUES (?, ?, ?)",
                 (attempt_id, payload.event_type, payload.details))
    new_count = attempt["malpractice_count"] + 1
    conn.execute("UPDATE test_attempts SET malpractice_count = ?, malpractice_flag = 1 WHERE attempt_id = ?", (new_count, attempt_id))
    conn.commit()
    conn.close()
    return {"status": "ok", "malpractice_count": new_count}


@app.post("/tests")
def create_test(payload: TestCreateRequest, request: Request):
    user = _get_current_user(request)
    _require_superuser(user)
    _ensure_tests_tables()
    conn = get_connection()
    cur = conn.execute(
        "INSERT INTO tests (title, description, duration_minutes, created_by) VALUES (?, ?, ?, ?)",
        (payload.title, payload.description, payload.duration_minutes, user["user"]),
    )
    test_id = cur.lastrowid
    for i, q in enumerate(payload.questions):
        conn.execute(
            """
            INSERT INTO test_questions
            (test_id, question_order, question_text, option_a, option_b, option_c, option_d, correct_answer, points)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (test_id, i + 1, q.question_text, q.option_a, q.option_b, q.option_c, q.option_d, q.correct_answer, 1),
        )
    for sid in payload.assigned_students:
        conn.execute(
            "INSERT OR IGNORE INTO test_assignments (test_id, student_id) VALUES (?, ?)",
            (test_id, sid),
        )
    conn.commit()
    conn.close()
    _log_activity(user, "test_created", f"Created test #{test_id} ({payload.title})", {
        "test_id": test_id,
        "title": payload.title,
    })
    return {"status": "ok", "test_id": test_id}


@app.post("/notifications/{notification_id}/read")
def mark_notification_read(notification_id: int, request: Request):
    user = _get_current_user(request)
    conn = get_connection()
    conn.execute(
        """
        INSERT OR IGNORE INTO notification_reads (notification_id, user_id)
        VALUES (?, ?)
        """,
        (notification_id, user["user"]),
    )
    conn.commit()
    conn.close()
    return {"status": "ok"}


@app.get("/feed")
def dashboard_feed(request: Request):
    user = _get_current_user(request)
    conn = get_connection()

    if user and not _is_superuser(user):
        fees = conn.execute(
            """
            SELECT COALESCE(SUM(amount_total),0) AS total,
                   COALESCE(SUM(amount_paid),0) AS paid,
                   COUNT(*) AS transactions
            FROM fees
            WHERE student_id = ?
            """,
            (user["user"],),
        ).fetchone()
    else:
        fees = conn.execute(
            """
            SELECT COALESCE(SUM(amount_total),0) AS total,
                   COALESCE(SUM(amount_paid),0) AS paid,
                   COUNT(*) AS transactions
            FROM fees
            """
        ).fetchone()

    announcements = conn.execute(
        "SELECT announcement_id, title, message, created_at FROM announcements ORDER BY announcement_id DESC LIMIT 5"
    ).fetchall()

    uid = user["user"]
    notifications = conn.execute(
        """
        SELECT notification_id, title, message, level, created_at
        FROM notifications
        WHERE target_user IS NULL OR target_user = ?
        ORDER BY notification_id DESC
        LIMIT 5
        """,
        (uid,),
    ).fetchall()

    interviews = conn.execute(
        "SELECT interview_id, airline_name, interview_date FROM interview_stats ORDER BY interview_date DESC LIMIT 5"
    ).fetchall()

    conn.close()

    total = float(fees["total"])
    paid = float(fees["paid"])
    return {
        "fees": {
            "total": total,
            "due": total - paid,
            "transactions": int(fees["transactions"]),
        },
        "announcements": [dict(r) for r in announcements],
        "notifications": [dict(r) for r in notifications],
        "interviews": [dict(r) for r in interviews],
    }

@app.get("/reports/summary")
def reports_summary(request: Request):
    _require_superuser(_get_current_user(request))
    conn = get_connection()
    student_count = conn.execute("SELECT COUNT(*) AS c FROM students").fetchone()["c"]
    student_rows = conn.execute("SELECT student_id FROM students").fetchall()
    fees_total = 0.0
    fees_paid = 0.0
    for row in student_rows:
        info = _student_financials(conn, row["student_id"])
        if not info:
            continue
        fees_total += float(info["total"])
        fees_paid += float(info["paid"])
    attendance_counts = conn.execute("""
        SELECT
            COALESCE(SUM(CASE WHEN attendance_status = 'Present' THEN 1 ELSE 0 END), 0) AS present,
            COALESCE(SUM(CASE WHEN attendance_status = 'Absent' THEN 1 ELSE 0 END), 0) AS absent
        FROM attendance
    """).fetchone()
    conn.close()
    return {
        "students": student_count,
        "fees_total": fees_total,
        "fees_paid": fees_paid,
        "fees_balance": fees_total - fees_paid,
        "attendance_present": attendance_counts["present"],
        "attendance_absent": attendance_counts["absent"],
    }
